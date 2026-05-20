import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

# Configurações de Design
HEADER_BG = "1F2937"  # Slate-900 (Estilo ZEHLA)
HEADER_TEXT = "FFFFFF"
ACCENT_ORANGE = "F97316"
TOTAL_BG = "F3F4F6"
BORDER_COLOR = "D1D5DB"

def apply_header_style(cell):
    cell.font = Font(name="Arial", size=10, bold=True, color=HEADER_TEXT)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

def apply_data_style(cell, row_idx, align="left"):
    cell.font = Font(name="Arial", size=10)
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
    else:
        cell.fill = PatternFill("solid", fgColor="F9FAFB")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

def create_v2_xlsx(output_path):
    wb = Workbook()
    
    # --- ABA 1: CALCULADORA DE ROI & PROJEÇÃO ---
    ws_roi = wb.active
    ws_roi.title = "Calculadora de ROI"
    
    ws_roi["B2"] = "ZEHLA SMARTHOTEL - CALCULADORA DE VIABILIDADE"
    ws_roi["B2"].font = Font(size=14, bold=True, color=HEADER_BG)
    
    # Inputs (Células que o usuário pode mudar)
    inputs = [
        ("Base de Leads", 10000, "B5"),
        ("Taxa de Conversão (%)", 0.02, "B6"),
        ("Churn Rate Mensal (%)", 0.05, "B7"),
        ("Ticket Médio (Calculado)", None, "B8"),
    ]
    
    ws_roi["B4"] = "VARIÁVEIS DE ENTRADA (MUDE OS VALORES)"
    ws_roi["B4"].font = Font(bold=True)
    
    for i, (label, val, pos) in enumerate(inputs):
        row = 5 + i
        ws_roi.cell(row=row, column=2, value=label).font = Font(bold=True)
        if val is not None:
            ws_roi.cell(row=row, column=3, value=val)
    
    ws_roi["C6"].number_format = "0.0%"
    ws_roi["C7"].number_format = "0.0%"
    
    # Matriz de Preços
    ws_roi["E4"] = "MATRIZ DE PLANOS"
    ws_roi["E4"].font = Font(bold=True)
    headers = ["Plano", "Preço (R$)", "Mix (%)", "Faturamento"]
    for i, h in enumerate(headers):
        cell = ws_roi.cell(row=5, column=5 + i, value=h)
        apply_header_style(cell)
    
    plans = [
        ("LITE", 248, 0.50),
        ("PRO", 448, 0.30),
        ("MAX", 798, 0.20)
    ]
    
    for i, (name, price, mix) in enumerate(plans):
        row = 6 + i
        ws_roi.cell(row=row, column=5, value=name)
        ws_roi.cell(row=row, column=6, value=price).number_format = '"R$" #,##0'
        ws_roi.cell(row=row, column=7, value=mix).number_format = "0%"
        # Faturamento por plano: (Leads * Conversão) * Mix * Preço
        ws_roi.cell(row=row, column=8, value=f"=($C$5 * $C$6) * G{row} * F{row}").number_format = '"R$" #,##0'
        for col in range(5, 9):
            apply_data_style(ws_roi.cell(row=row, column=col), i)
    
    # Ticket Médio Formula
    ws_roi["C8"] = "=SUM(H6:H8) / ($C$5 * $C$6)"
    ws_roi["C8"].number_format = '"R$" #,##0'
    
    # RESULTADOS
    ws_roi["B11"] = "RESULTADOS PROJETADOS"
    ws_roi["B11"].font = Font(bold=True, size=12)
    
    results = [
        ("Total de Clientes Ativos", "=$C$5 * $C$6", "clientes"),
        ("MRR (Faturamento Mensal)", "=SUM(H6:H8)", "R$"),
        ("ARR (Faturamento Anual)", "=C13 * 12", "R$"),
        ("Custos Operacionais Fixos", 442.13, "R$"),
        ("EBITDA (Lucro Operacional)", "=C13 - C15", "R$"),
        ("Margem Líquida (%)", "=C16 / C13", "%")
    ]
    
    for i, (label, formula, unit) in enumerate(results):
        row = 12 + i
        ws_roi.cell(row=row, column=2, value=label).font = Font(bold=True)
        cell = ws_roi.cell(row=row, column=3, value=formula)
        if unit == "R$":
            cell.number_format = '"R$" #,##0.00'
        elif unit == "%":
            cell.number_format = "0.0%"
        else:
            cell.number_format = "#,##0"
        
        # Style
        if "Lucro" in label:
            cell.font = Font(bold=True, color="16A34A")
    
    # --- ABA 2: CUSTOS DETALHADOS ---
    ws_costs = wb.create_sheet("Custos & Infra")
    ws_costs["B2"] = "DETALHAMENTO DE CUSTOS OPERACIONAIS"
    ws_costs["B2"].font = Font(size=12, bold=True)
    
    headers = ["Categoria", "Item", "Valor (R$)", "Frequência", "Subtotal (Anual)"]
    for i, h in enumerate(headers):
        cell = ws_costs.cell(row=4, column=2 + i, value=h)
        apply_header_style(cell)
        
    items = [
        ("Infra", "Hostinger VPS KVM 4", 149.90, "Mensal"),
        ("Infra", "PostgreSQL Managed", 49.90, "Mensal"),
        ("Infra", "Redis Cache", 29.90, "Mensal"),
        ("Infra", "Domínio .com.br", 3.33, "Mensal"),
        ("Ferramentas", "Z-API WhatsApp", 89.90, "Mensal"),
        ("Ferramentas", "Kimi 2.6 API", 199.00, "Mensal (Est.)"),
        ("Ferramentas", "ElevenLabs Voice", 99.00, "Mensal")
    ]
    
    for i, (cat, item, val, freq) in enumerate(items):
        row = 5 + i
        ws_costs.cell(row=row, column=2, value=cat)
        ws_costs.cell(row=row, column=3, value=item)
        ws_costs.cell(row=row, column=4, value=val).number_format = '"R$" #,##0.00'
        ws_costs.cell(row=row, column=5, value=freq)
        ws_costs.cell(row=row, column=6, value=f"=D{row} * 12").number_format = '"R$" #,##0.00'
        for col in range(2, 7):
            apply_data_style(ws_costs.cell(row=row, column=col), i)
            
    # Totais
    total_row = 5 + len(items)
    ws_costs.cell(row=total_row, column=3, value="TOTAL MENSAL").font = Font(bold=True)
    ws_costs.cell(row=total_row, column=4, value=f"=SUM(D5:D{total_row-1})").number_format = '"R$" #,##0.00'
    ws_costs.cell(row=total_row, column=4).font = Font(bold=True)
    
    # --- ABA 3: CRONOGRAMA ---
    ws_cron = wb.create_sheet("Cronograma 2026")
    # (Simplified for now, focusing on formulas)
    ws_cron["B2"] = "CRONOGRAMA DE LANÇAMENTO (DIAS ÚTEIS)"
    
    # Auto-adjust columns
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    print(f"Planilha V2 salva em: {output_path}")

if __name__ == "__main__":
    create_v2_xlsx("/Users/marciocau/Downloads/ENTREGÁVEIS ZEHLA FULL STACK/ZEHLA_Plano_Master_V2_Fórmulas.xlsx")
