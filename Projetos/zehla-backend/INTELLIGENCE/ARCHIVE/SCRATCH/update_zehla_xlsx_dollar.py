import openpyxl
from openpyxl.styles import Font, PatternFill

def update_xlsx_dynamic_dollar(file_path):
    wb = openpyxl.load_workbook(file_path)
    
    # Aba ROI - Adicionar Cotação do Dólar
    ws_roi = wb["Calculadora de ROI"]
    
    ws_roi["B3"] = "COTAÇÃO USD/BRL (HOJE)"
    ws_roi["B3"].font = Font(bold=True, color="16A34A")
    ws_roi["C3"] = 5.85 # Valor exemplo que o usuário pode mudar
    ws_roi["C3"].number_format = '"R$" #,##0.00'
    
    # Atualizar Custos Operacionais Fixos (C15) para usar conversão se necessário
    # Mas é melhor fazer na aba de Custos primeiro
    
    # Aba Custos
    if "Custos & Infra" in wb.sheetnames:
        ws_costs = wb["Custos & Infra"]
        
        # Inserir a cotação também na aba de custos para referência
        ws_costs["H4"] = "Cotação USD"
        ws_costs["I4"] = "='Calculadora de ROI'!$C$3"
        ws_costs["I4"].number_format = '"R$" #,##0.00'
        
        # Atualizar itens que são em Dólar
        items_in_usd = {
            "Amazon SES (E-mail)": 1.00, # USD
            "Kimi 2.6 API": 35.00, # USD (Exemplo)
            "ElevenLabs Voice": 18.00 # USD (Exemplo)
        }
        
        for row in range(5, 20):
            item_name = ws_costs.cell(row=row, column=3).value
            if item_name in items_in_usd:
                usd_val = items_in_usd[item_name]
                ws_costs.cell(row=row, column=4, value="=$I$4 * {}".format(usd_val))
                ws_costs.cell(row=row, column=4).number_format = '"R$" #,##0.00'
                ws_costs.cell(row=row, column=7, value=f"Valor Original: ${usd_val} USD")
                
        # Recalcular Total Mensal (Célula D12 por exemplo)
        # Vamos assumir que o total está na linha após os itens
        total_row = 15 # Ajustar conforme necessário
        ws_costs.cell(row=total_row, column=3, value="TOTAL MENSAL (BR)").font = Font(bold=True)
        ws_costs.cell(row=total_row, column=4, value="=SUM(D5:D{})".format(total_row-1)).number_format = '"R$" #,##0.00'

    wb.save(file_path)
    print(f"Planilha atualizada com Dólar Dinâmico: {file_path}")

if __name__ == "__main__":
    update_xlsx_dynamic_dollar("/Users/marciocau/Downloads/ENTREGÁVEIS ZEHLA FULL STACK/ZEHLA_Plano_Master_V2_Fórmulas.xlsx")
