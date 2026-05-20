import openpyxl

def update_xlsx_costs(file_path):
    wb = openpyxl.load_workbook(file_path)
    
    # Aba ROI
    if "Calculadora de ROI" in wb.sheetnames:
        ws_roi = wb["Calculadora de ROI"]
        # Atualizar custos operacionais fixos (C15)
        # VPS (150) + PG (50) + Redis (30) + Z-API (90) + SES (5) + Listmonk (0) = ~325
        ws_roi["C15"] = 325.00
    
    # Aba Custos
    if "Custos & Infra" in wb.sheetnames:
        ws_costs = wb["Custos & Infra"]
        # Substituir ElevenLabs ou outras ferramentas de email por Amazon SES
        found = False
        for row in range(5, 20):
            item = ws_costs.cell(row=row, column=3).value
            if item and ("ElevenLabs" in item or "Email" in item):
                ws_costs.cell(row=row, column=2, value="Ferramentas")
                ws_costs.cell(row=row, column=3, value="Amazon SES (E-mail)")
                ws_costs.cell(row=row, column=4, value=5.50) # ~$1.00 + IOF
                ws_costs.cell(row=row, column=6, value="=D{}*12".format(row))
                found = True
                break
        
        if not found:
            # Adicionar nova linha
            row = 12
            ws_costs.cell(row=row, column=2, value="Ferramentas")
            ws_costs.cell(row=row, column=3, value="Amazon SES (E-mail)")
            ws_costs.cell(row=row, column=4, value=5.50)
            ws_costs.cell(row=row, column=5, value="Mensal (Uso)")
            ws_costs.cell(row=row, column=6, value="=D{}*12".format(row))
            
        # Adicionar Listmonk com custo zero
        row = 13
        ws_costs.cell(row=row, column=2, value="Ferramentas")
        ws_costs.cell(row=row, column=3, value="Listmonk Engine")
        ws_costs.cell(row=row, column=4, value=0.00)
        ws_costs.cell(row=row, column=5, value="Self-hosted")
        ws_costs.cell(row=row, column=6, value=0.00)

    wb.save(file_path)
    print(f"Planilha atualizada com custos de E-mail: {file_path}")

if __name__ == "__main__":
    update_xlsx_costs("/Users/marciocau/Downloads/ENTREGÁVEIS ZEHLA FULL STACK/ZEHLA_Plano_Master_V2_Fórmulas.xlsx")
