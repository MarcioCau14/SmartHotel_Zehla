import openpyxl
import sys
import json

def append_lead():
    if len(sys.argv) < 2:
        print("Error: Missing JSON file path")
        sys.exit(1)
        
    try:
        with open(sys.argv[1], 'r', encoding='utf-8') as f:
            payload = json.load(f)
        
        file_path = '/Users/marciocau/Downloads/POUSADAS_PDR (1).xlsx'
        wb = openpyxl.load_workbook(file_path)
        
        sheet_name = 'LEADS_VIA_LANDPAGE'
        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(sheet_name)
            headers = ['Pousada', 'e-mail', 'Whatsapp', 'Quant. quartos', 'Local', 'Cidade', 'UF', 'VALORES', 'Qualificação', 'Validação Contato', 'Comportamento de Compra', 'Sinais de Intenção']
            sheet.append(headers)
        else:
            sheet = wb[sheet_name]
            
        row = [
            payload.get('pousada', ''),
            payload.get('email', ''),
            payload.get('whatsapp', ''),
            payload.get('quant_quartos', ''),
            payload.get('local', ''),
            payload.get('cidade', ''),
            payload.get('uf', ''),
            payload.get('valores', ''),
            payload.get('qualificacao', ''),
            payload.get('validacao_contato', ''),
            payload.get('comportamento', ''),
            payload.get('sinais', '')
        ]
        
        sheet.append(row)
        wb.save(file_path)
        print("Success: Lead appended successfully")
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    append_lead()
