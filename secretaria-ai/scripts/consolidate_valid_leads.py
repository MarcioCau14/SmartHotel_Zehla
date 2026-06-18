import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==============================================================================
# CONFIGURAÇÕES E PASTAS
# ==============================================================================
FOLDER_PATH = "/Users/marciocau/Downloads/PLANILHAS_MARKETING_BR_"
OUTPUT_FILE_PATH = os.path.join(FOLDER_PATH, "CONTATOS_VALIDOS_CONSOLIDADO.xlsx")
SOURCE_FASE_VALIDADA = "/Users/marciocau/Downloads/POUSADAS_MARKETING_FASE_VALIDADA.xlsx"

# Estilos de design premium
HEADER_FILL = PatternFill(start_color="2A4B7C", end_color="2A4B7C", fill_type="solid") # Azul corporativo premium
HEADER_FONT = Font(color="FFFFFF", name="Segoe UI", size=11, bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROW_FONT = Font(name="Segoe UI", size=11)
BORDER_THIN = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# ==============================================================================
# PIPELINE DE CONSOLIDAÇÃO
# ==============================================================================

def main():
    print("🧠 [SECRETARIA-IA] Iniciando Consolidação de Contatos Válidos por Cidades...")
    
    # 1. Listar arquivos _VALIDADA.xlsx na pasta
    if not os.path.exists(FOLDER_PATH):
        print(f"❌ Pasta não existe: {FOLDER_PATH}")
        return
        
    xlsx_files = []
    for f in os.listdir(FOLDER_PATH):
        if f.endswith("_VALIDADA.xlsx") and f != "CONTATOS_VALIDOS_CONSOLIDADO.xlsx":
            xlsx_files.append(os.path.join(FOLDER_PATH, f))
            
    if os.path.exists(SOURCE_FASE_VALIDADA):
        xlsx_files.append(SOURCE_FASE_VALIDADA)
        
    print(f"📊 Mapeados {len(xlsx_files)} arquivos validados para extração de contatos.")
    
    # Dicionário de leads válidos
    # Chave: (Pousada, E-mail, Whatsapp) para remoção absoluta de duplicidades
    all_valid_leads = {}
    
    for file_path in xlsx_files:
        print(f"📂 Extraindo contatos válidos de: {os.path.basename(file_path)}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            
            col_indices = {}
            for col_idx, h in enumerate(headers, 1):
                h_lower = h.lower()
                if "pousada" in h_lower or "nome" in h_lower or "propriedade" in h_lower:
                    col_indices["Pousada"] = col_idx
                elif "e-mail" in h_lower or "email" in h_lower:
                    col_indices["Email"] = col_idx
                elif "whats" in h_lower:
                    col_indices["Whatsapp"] = col_idx
                elif "quarto" in h_lower:
                    col_indices["Qtd Quartos"] = col_idx
                elif "local" in h_lower or "praia" in h_lower:
                    col_indices["Local"] = col_idx
                elif "cidade" in h_lower:
                    col_indices["Cidade"] = col_idx
                elif "uf" in h_lower:
                    col_indices["UF"] = col_idx
                elif "valor" in h_lower:
                    col_indices["Valores Estimados"] = col_idx
                elif "qualificação" in h_lower or "qualificacao" in h_lower:
                    col_indices["Qualificação"] = col_idx
                elif "validação" in h_lower or "validacao" in h_lower:
                    col_indices["Validação"] = col_idx
                elif "comportamento" in h_lower:
                    col_indices["Comportamento de Compra"] = col_idx
                elif "sinal" in h_lower or "intenção" in h_lower or "intencao" in h_lower:
                    col_indices["Sinais de Intenção"] = col_idx
                elif "social" in h_lower or "rede" in h_lower:
                    col_indices["Redes Sociais"] = col_idx
                elif "latitude" in h_lower:
                    col_indices["LATITUDE"] = col_idx
                elif "longitude" in h_lower:
                    col_indices["LONGITUDE"] = col_idx
                elif "score qual" in h_lower:
                    col_indices["Score Qual."] = col_idx
                elif "score valid" in h_lower:
                    col_indices["Score Valid."] = col_idx
            
            if "Pousada" not in col_indices:
                continue
                
            for row in range(2, sheet.max_row + 1):
                pousada_val = sheet.cell(row=row, column=col_indices["Pousada"]).value
                if not pousada_val:
                    continue
                    
                pousada_name = str(pousada_val).strip()
                
                email_val = str(sheet.cell(row=row, column=col_indices["Email"]).value).strip() if "Email" in col_indices and sheet.cell(row=row, column=col_indices["Email"]).value is not None else ""
                whatsapp_val = str(sheet.cell(row=row, column=col_indices["Whatsapp"]).value).strip() if "Whatsapp" in col_indices and sheet.cell(row=row, column=col_indices["Whatsapp"]).value is not None else ""
                uf_val = str(sheet.cell(row=row, column=col_indices["UF"]).value).strip() if "UF" in col_indices and sheet.cell(row=row, column=col_indices["UF"]).value is not None else "SC"
                cidade_val = str(sheet.cell(row=row, column=col_indices["Cidade"]).value).strip() if "Cidade" in col_indices and sheet.cell(row=row, column=col_indices["Cidade"]).value is not None else "Não Informada"
                local_val = str(sheet.cell(row=row, column=col_indices["Local"]).value).strip() if "Local" in col_indices and sheet.cell(row=row, column=col_indices["Local"]).value is not None else ""
                quartos_val = sheet.cell(row=row, column=col_indices["Qtd Quartos"]).value if "Qtd Quartos" in col_indices else ""
                valores_val = sheet.cell(row=row, column=col_indices["Valores Estimados"]).value if "Valores Estimados" in col_indices else ""
                qual_val = sheet.cell(row=row, column=col_indices["Qualificação"]).value if "Qualificação" in col_indices else ""
                val_status = sheet.cell(row=row, column=col_indices["Validação"]).value if "Validação" in col_indices else ""
                comp_val = sheet.cell(row=row, column=col_indices["Comportamento de Compra"]).value if "Comportamento de Compra" in col_indices else ""
                sinais_val = sheet.cell(row=row, column=col_indices["Sinais de Intenção"]).value if "Sinais de Intenção" in col_indices else ""
                redes_val = sheet.cell(row=row, column=col_indices["Redes Sociais"]).value if "Redes Sociais" in col_indices else ""
                lat_val = sheet.cell(row=row, column=col_indices["LATITUDE"]).value if "LATITUDE" in col_indices else None
                lon_val = sheet.cell(row=row, column=col_indices["LONGITUDE"]).value if "LONGITUDE" in col_indices else None
                score_q = sheet.cell(row=row, column=col_indices["Score Qual."]).value if "Score Qual." in col_indices else 0
                score_v = sheet.cell(row=row, column=col_indices["Score Valid."]).value if "Score Valid." in col_indices else 0
                
                # Validação lógica do lead (somente contatos realmente válidos)
                # O e-mail e whatsapp precisam estar preenchidos e válidos de acordo com o status
                email_ok = email_val and email_val != "-" and email_val != "nan" and "@" in email_val
                whats_ok = whatsapp_val and whatsapp_val != "-" and whatsapp_val != "nan"
                
                if val_status:
                    val_str = str(val_status).lower()
                    if "domínio sem registros mx" in val_str or "sintaxe de e-mail inválida" in val_str or "e-mail ausente" in val_str:
                        email_ok = False
                    if "whatsapp ausente" in val_str or "número inválido" in val_str or "tamanho de número inválido" in val_str:
                        whats_ok = False
                
                # Somente adiciona se ambos forem de fato válidos (higienizados)
                if email_ok and whats_ok:
                    lead_key = (pousada_name.lower().strip(), email_val.lower().strip(), re.sub(r"\D", "", whatsapp_val))
                    
                    # Guarda os dados completos
                    all_valid_leads[lead_key] = {
                        "pousada": pousada_name,
                        "email": email_val,
                        "whatsapp": whatsapp_val,
                        "quartos": quartos_val,
                        "local": local_val,
                        "cidade": cidade_val,
                        "uf": uf_val,
                        "valores": valores_val,
                        "qualificacao": qual_val,
                        "validacao": val_status,
                        "comportamento": comp_val,
                        "sinais": sinais_val,
                        "redes": redes_val,
                        "lat": lat_val,
                        "lon": lon_val,
                        "score_qual": score_q,
                        "score_valid": score_v
                    }
                    
        except Exception as e:
            print(f"❌ Erro ao extrair de {file_path}: {e}")
            
    print(f"📊 Total de leads válidos e exclusivos carregados: {len(all_valid_leads)}")
    
    # 2. Agrupar por Cidades
    leads_by_city = {}
    for lead in all_valid_leads.values():
        city_raw = lead["cidade"].strip()
        # Normaliza o nome da cidade para evitar duplicatas por caixa (ex: "Imbituba" e "IMBITUBA")
        city_name = city_raw.title()
        
        if city_name not in leads_by_city:
            leads_by_city[city_name] = []
        leads_by_city[city_name].append(lead)
        
    print(f"🌆 Encontradas {len(leads_by_city)} cidades distintas com contatos válidos.")
    
    # 3. Criar Novo Workbook Excel Consolidado
    wb_out = openpyxl.Workbook()
    # Remove a aba padrão
    default_sheet = wb_out.active
    wb_out.remove(default_sheet)
    
    # Definir cabeçalhos padrão (todas as 18 colunas)
    columns_headers = [
        '#', 'Pousada', 'E-mail', 'Whatsapp', 'Qtd Quartos', 'Local / Praia', 'Cidade', 'UF', 'Valores Estimados',
        'Qualificação', 'Validação', 'Comportamento de Compra', 'Sinais de Intenção', 'Redes Sociais', 'LATITUDE', 'LONGITUDE', 'Score Qual.', 'Score Valid.'
    ]
    
    # Mapeamento chave-valor dos dados
    data_keys = [
        "pousada", "email", "whatsapp", "quartos", "local", "cidade", "uf", "valores",
        "qualificacao", "validacao", "comportamento", "sinais", "redes", "lat", "lon", "score_qual", "score_valid"
    ]
    
    # Primeiro criamos a aba de "TODOS OS CONTATOS" para facilitar visualização geral
    all_leads_sorted = []
    for city_name, leads in sorted(leads_by_city.items()):
        all_leads_sorted.extend(leads)
        
    # Salva todos em uma aba principal
    create_styled_sheet(wb_out, "Todos_Validos", columns_headers, all_leads_sorted, data_keys)
    
    # Cria uma aba específica para cada cidade
    # Limita o nome da aba ao máximo de 30 caracteres permitido pelo Excel
    for city_name, leads in sorted(leads_by_city.items()):
        # Sanitiza o nome da aba
        sheet_title = re.sub(r"[\\/*?:\[\]]", "", city_name)[:30]
        if not sheet_title:
            sheet_title = "Outros"
        create_styled_sheet(wb_out, sheet_title, columns_headers, leads, data_keys)
        
    # Salva o arquivo final
    wb_out.save(OUTPUT_FILE_PATH)
    print(f"\n🎉 [SUCESSO-TOTAL] Planilha consolidada de contatos válidos por cidades criada em: {OUTPUT_FILE_PATH}")
    
def create_styled_sheet(wb, title, headers, leads, keys):
    """
    Cria uma aba estilizada de alto padrão no workbook com auto-largura e bordas.
    """
    sheet = wb.create_sheet(title=title)
    sheet.views.sheetView[0].showGridLines = True
    
    # Escreve cabeçalhos
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
        
    sheet.row_dimensions[1].height = 28
    
    # Escreve as linhas de leads
    for row_idx, lead in enumerate(leads, 2):
        sheet.row_dimensions[row_idx].height = 20
        
        # Coluna sequencial #
        num_cell = sheet.cell(row=row_idx, column=1)
        num_cell.value = row_idx - 1
        num_cell.font = ROW_FONT
        num_cell.alignment = Alignment(horizontal="center")
        num_cell.border = BORDER_THIN
        
        for col_idx, key in enumerate(keys, 2):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = lead[key]
            cell.font = ROW_FONT
            cell.border = BORDER_THIN
            
            # Alinhamento específico por tipo de dado
            if key in ["whatsapp", "quartos", "score_qual", "score_valid", "lat", "lon", "uf"]:
                cell.alignment = Alignment(horizontal="center")
            elif key in ["email"]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-ajuste de largura de colunas (estilo premium)
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Define a largura com espaçamento extra de segurança
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 45)

if __name__ == "__main__":
    main()
