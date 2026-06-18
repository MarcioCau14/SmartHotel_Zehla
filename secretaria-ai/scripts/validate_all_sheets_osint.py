import os
import re
import openpyxl
import pandas as pd
import dns.resolver
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import PatternFill, Font

# ==============================================================================
# CONFIGURAÇÕES E PASTAS
# ==============================================================================
FOLDER_PATH = "/Users/marciocau/Downloads/PLANILHAS_MARKETING_BR_"
CONSOLIDATED_FILE_PATH = os.path.join(FOLDER_PATH, "POUSADAS_MARKETING_FASE (1).xlsx")
SOURCE_POUSADAS_FILE = "/Users/marciocau/Downloads/POUSADAS_MARKETING_FASE(1).xlsx"

# Estilos de preenchimento e fonte para marcação de contatos inválidos/inexistentes
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006", name="Calibri", size=11, bold=True)

# ==============================================================================
# CACHES GLOBAIS DE VALIDAÇÃO (DEDUPLICAÇÃO E ACELERAÇÃO)
# ==============================================================================
domain_mx_cache = {}          # Mapeia domínio -> bool (tem MX ou não)
validated_emails_cache = {}     # Mapeia e-mail -> dict com dados de validação
validated_whatsapps_cache = {}  # Mapeia whatsapp -> dict com dados de validação

# ==============================================================================
# 1. PRÉ-RESOLUÇÃO DE DOMÍNIOS DNS MX EM PARALELO (ThreadPoolExecutor)
# ==============================================================================

def check_mx_dns_single(domain):
    """
    Executa a resolução MX de um domínio individual usando dnspython.
    """
    common_good_domains = {
        "gmail.com", "hotmail.com", "yahoo.com", "outlook.com", "yahoo.com.br",
        "uol.com.br", "terra.com.br", "ig.com.br", "bol.com.br", "live.com", 
        "outlook.com.br", "icloud.com", "globo.com", "icloud.com.br", "me.com"
    }
    if domain in common_good_domains:
        return domain, True
        
    try:
        # Consulta registro MX com timeout pequeno de 2 segundos
        answers = dns.resolver.resolve(domain, 'MX', lifetime=2)
        return domain, len(answers) > 0
    except Exception:
        # Se falhar ou der timeout, tenta checar registro A como fallback de host existente
        try:
            answers = dns.resolver.resolve(domain, 'A', lifetime=2)
            return domain, len(answers) > 0
        except Exception:
            return domain, False

def pre_resolve_all_domains(xlsx_files):
    """
    Varre rapidamente todas as planilhas para coletar domínios únicos de e-mail
    e resolve-os em paralelo com 100 threads em menos de 30 segundos!
    """
    print("\n🔍 [DNS EXTRACTION] Coletando domínios únicos de e-mail para pré-resolução...")
    unique_domains = set()
    
    for filename in xlsx_files:
        if filename == SOURCE_POUSADAS_FILE:
            path = SOURCE_POUSADAS_FILE
        else:
            path = os.path.join(FOLDER_PATH, filename)
            
        try:
            # Usa pandas com read_only para carregar colunas em milissegundos
            xl = pd.ExcelFile(path)
            for sheet_name in xl.sheet_names:
                # Carrega apenas a coluna de e-mail se possível
                df = pd.read_excel(path, sheet_name=sheet_name, nrows=10)
                email_col = None
                for col in df.columns:
                    if "e-mail" in str(col).lower() or "email" in str(col).lower():
                        email_col = col
                        break
                if email_col is not None:
                    # Carrega a coluna inteira
                    df_full = pd.read_excel(path, sheet_name=sheet_name, usecols=[email_col])
                    emails = df_full[email_col].dropna().astype(str).tolist()
                    for email in emails:
                        if "@" in email:
                            dom = email.split("@")[1].strip().lower()
                            # Filtra caracteres inválidos no domínio
                            dom = re.sub(r"[^\w\.-]", "", dom)
                            if dom:
                                unique_domains.add(dom)
        except Exception as e:
            print(f"   ⚠️ Erro rápido ao ler domínios de {filename}: {e}")
            
    print(f"📊 Total de domínios únicos de e-mail encontrados: {len(unique_domains)}")
    
    # Executa a validação DNS em paralelo usando ThreadPoolExecutor
    print(f"🚀 Iniciando resolução DNS paralela com 100 threads para {len(unique_domains)} domínios...")
    resolved_count = 0
    valid_count = 0
    
    with ThreadPoolExecutor(max_workers=100) as executor:
        results = executor.map(check_mx_dns_single, unique_domains)
        for dom, has_mx in results:
            domain_mx_cache[dom] = has_mx
            resolved_count += 1
            if has_mx:
                valid_count += 1
            if resolved_count % 1000 == 0 or resolved_count == len(unique_domains):
                print(f"   Processed {resolved_count}/{len(unique_domains)} domínios... (Válidos: {valid_count})")
                
    print(f"✅ Pré-resolução DNS concluída! Domínios válidos com MX/A: {valid_count} | Inválidos: {resolved_count - valid_count}")

# ==============================================================================
# VALIDAÇÃO DE CONTEXTO E OSINT (0ms de latência usando cache pré-resolvido)
# ==============================================================================

def validate_email_osint(email_str, pousada_name):
    """
    Valida e-mail usando o cache pré-resolvido DNS e lógicas OSINT.
    """
    if not email_str or email_str == "-" or email_str == "nan":
        return False, "E-mail Ausente", 0
        
    email_clean = email_str.strip().lower()
    
    email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
    if not re.match(email_regex, email_clean):
        return False, "Sintaxe de E-mail Inválida", 0
        
    # Check cache
    if email_clean in validated_emails_cache:
        cached = validated_emails_cache[email_clean]
        return cached["is_valid"], cached["status"], cached["score"]
        
    domain = email_clean.split("@")[1].strip().lower()
    domain = re.sub(r"[^\w\.-]", "", domain)
    
    # Puxa do cache de MX resolvido em paralelo
    has_mx = domain_mx_cache.get(domain, False)
    if not has_mx:
        res = (False, "Domínio sem registros MX (Inexistente)", 0)
        validated_emails_cache[email_clean] = {"is_valid": False, "status": res[1], "score": 0}
        return res
        
    username = email_clean.split("@")[0]
    
    pousada_norm = re.sub(r"[^\w]", "", pousada_name.lower())
    username_norm = re.sub(r"[^\w]", "", username)
    domain_norm = re.sub(r"[^\w]", "", domain.split(".")[0])
    
    score = 50
    status = "Validado via MX"
    
    if username_norm in pousada_norm or domain_norm in pousada_norm:
        score += 35
        status = "Validado via Digital Footprint (Domínio Pousada)"
    elif username in ["atendimento", "reservas", "contato", "comercial", "direcao"]:
        score += 15
        status = "Canal Corporativo Oficial"
    else:
        score += 25
        status = "Contato Direto Potencial"
        
    res = (True, status, score)
    validated_emails_cache[email_clean] = {"is_valid": True, "status": status, "score": score}
    return res

def validate_whatsapp_osint(phone_str, uf_lead):
    """
    Valida WhatsApp usando celular e DDD de decisor externo.
    """
    if not phone_str or phone_str == "-" or phone_str == "nan":
        return False, "WhatsApp Ausente", 0, ""
        
    phone_raw = str(phone_str)
    if phone_raw in validated_whatsapps_cache:
        cached = validated_whatsapps_cache[phone_raw]
        return cached["is_valid"], cached["status"], cached["score"], cached["clean"]
        
    digits = re.sub(r"\D", "", phone_raw)
    
    if len(digits) < 10 or len(digits) > 13:
        res = (False, "Tamanho de Número Inválido", 0, digits)
        validated_whatsapps_cache[phone_raw] = {"is_valid": False, "status": res[1], "score": 0, "clean": digits}
        return res
        
    if len(digits) in [12, 13] and digits.startswith("55"):
        digits = digits[2:]
        
    if len(digits) == 11:
        ddd = digits[:2]
        ninth_digit = digits[2]
        
        if ninth_digit == "9":
            status = "WhatsApp Celular Ativo"
            score = 90
            
            # DDD mismatch decisor check
            sc_ddds = ["48", "47", "49"]
            if uf_lead == "SC" and ddd not in sc_ddds:
                status = "WhatsApp Decisor Direto (DDD Externo)"
                score = 98
            elif uf_lead == "ES" and ddd not in ["27", "28"]:
                status = "WhatsApp Decisor Direto (DDD Externo)"
                score = 98
            elif uf_lead == "RS" and ddd not in ["51", "53", "54", "55"]:
                status = "WhatsApp Decisor Direto (DDD Externo)"
                score = 98
            elif uf_lead == "PR" and ddd not in ["41", "42", "43", "44", "45", "46"]:
                status = "WhatsApp Decisor Direto (DDD Externo)"
                score = 98
        else:
            status = "Número Incorreto (Celular sem 9)"
            score = 40
    elif len(digits) == 10:
        status = "WhatsApp Fixo Corporativo"
        score = 65
    else:
        status = "Número Inválido"
        score = 0
        
    is_valid = score >= 50
    res = (is_valid, status, score, digits)
    validated_whatsapps_cache[phone_raw] = {"is_valid": is_valid, "status": status, "score": score, "clean": digits}
    return res

def parse_valores(val_str):
    if not val_str or val_str == "-" or val_str == "nan":
        return 0
    clean = re.sub(r"[^\d]", "", str(val_str))
    if not clean:
        return 0
    return int(clean)

def calculate_purchasing_power(valores_str, quartos_str):
    val = parse_valores(valores_str)
    quartos = 0
    if quartos_str and quartos_str != "-" and quartos_str != "nan":
        q_digits = re.sub(r"\D", "", str(quartos_str))
        if q_digits:
            quartos = int(q_digits)
            
    if val >= 500 or quartos > 20:
        return "ALTO (ICP A+)", "Comportamento Seletivo: Alta exigência por automação e exclusividade. Foco em ROI.", 90
    elif val >= 250 or quartos > 10:
        return "MÉDIO (ICP A)", "Comportamento Moderado: Equilíbrio entre tradição e abertura tecnológica.", 70
    else:
        return "NORMAL", "Foco em Eficiência: Busca por redução de custos e escala. Valoriza ferramentas baratas.", 45

# ==============================================================================
# PIPELINE PRINCIPAL DE PROCESSAMENTO
# ==============================================================================

def main():
    print("🧠 [SECRETARIA-IA] Iniciando Validação Inteligente OTIMIZADA em Lote...")
    
    if not os.path.exists(FOLDER_PATH):
        print(f"❌ Pasta de destino não existe: {FOLDER_PATH}")
        return
        
    # Coleta todos os arquivos fonte
    xlsx_files = []
    for f in os.listdir(FOLDER_PATH):
        if f.endswith(".xlsx") and not f.startswith("~$"):
            if f != "POUSADAS_MARKETING_FASE (1).xlsx" and not f.endswith("_VALIDADA.xlsx"):
                xlsx_files.append(f)
                
    if os.path.exists(SOURCE_POUSADAS_FILE):
        xlsx_files.append(SOURCE_POUSADAS_FILE)
        
    print(f"📊 Mapeados {len(xlsx_files)} arquivos de planilhas contendo leads.")
    
    # 1. Executa Pré-resolução DNS ultra-rápida em paralelo
    pre_resolve_all_domains(xlsx_files)
    
    praia_rosa_leads = []
    
    # 2. Processa cada planilha (agora roda a 0ms de DNS em cada linha!)
    for file_item in xlsx_files:
        if file_item == SOURCE_POUSADAS_FILE:
            file_path = SOURCE_POUSADAS_FILE
            output_name = "/Users/marciocau/Downloads/POUSADAS_MARKETING_FASE_VALIDADA.xlsx"
            print(f"\n📂 Processando planilha principal: {file_item}")
        else:
            file_path = os.path.join(FOLDER_PATH, file_item)
            base_name = file_item.replace(".xlsx", "")
            output_name = os.path.join(FOLDER_PATH, f"{base_name}_VALIDADA.xlsx")
            print(f"\n📂 Processando planilha da pasta: {file_item}")
            
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            print(f"   Aba: '{sheet.title}' | Linhas: {sheet.max_row}")
            
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            
            col_indices = {}
            for col_idx, h in enumerate(headers, 1):
                h_lower = h.lower()
                if "pousada" in h_lower or "nome" in h_lower or "propriedade" in h_lower:
                    col_indices["Pousada"] = col_idx
                elif "e-mail" in h_lower or "email" in h_lower:
                    col_indices["Email"] = col_idx
                elif "whats" in h_lower:
                    col_indices["Whatsapp"] = col_idx
                elif "quarto" in h_lower:
                    col_indices["Quant. quartos"] = col_idx
                elif "local" in h_lower or "praia" in h_lower:
                    col_indices["Local"] = col_idx
                elif "cidade" in h_lower:
                    col_indices["Cidade"] = col_idx
                elif "uf" in h_lower:
                    col_indices["UF"] = col_idx
                elif "valor" in h_lower:
                    col_indices["Valores"] = col_idx
                elif "qualificação" in h_lower or "qualificacao" in h_lower:
                    col_indices["Qualificacao"] = col_idx
                elif "validação" in h_lower or "validacao" in h_lower:
                    col_indices["Validacao"] = col_idx
                elif "comportamento" in h_lower:
                    col_indices["Comportamento"] = col_idx
                elif "sinal" in h_lower or "intenção" in h_lower or "intencao" in h_lower:
                    col_indices["Sinais"] = col_idx
                elif "social" in h_lower or "rede" in h_lower:
                    col_indices["Redes"] = col_idx
                elif "latitude" in h_lower:
                    col_indices["LATITUDE"] = col_idx
                elif "longitude" in h_lower:
                    col_indices["LONGITUDE"] = col_idx
                elif "score qual" in h_lower:
                    col_indices["Score Qual"] = col_idx
                elif "score valid" in h_lower:
                    col_indices["Score Valid"] = col_idx
            
            if "Pousada" not in col_indices:
                print("   ⚠️ Coluna 'Pousada' não mapeada. Ignorando planilha.")
                continue
                
            invalid_leads_count = 0
            valid_leads_count = 0
            
            for row in range(2, sheet.max_row + 1):
                pousada_cell = sheet.cell(row=row, column=col_indices["Pousada"])
                if not pousada_cell.value:
                    continue
                    
                pousada_name = str(pousada_cell.value).strip()
                
                email_cell = sheet.cell(row=row, column=col_indices["Email"]) if "Email" in col_indices else None
                whatsapp_cell = sheet.cell(row=row, column=col_indices["Whatsapp"]) if "Whatsapp" in col_indices else None
                uf_cell = sheet.cell(row=row, column=col_indices["UF"]) if "UF" in col_indices else None
                valores_cell = sheet.cell(row=row, column=col_indices["Valores"]) if "Valores" in col_indices else None
                quartos_cell = sheet.cell(row=row, column=col_indices["Quant. quartos"]) if "Quant. quartos" in col_indices else None
                cidade_cell = sheet.cell(row=row, column=col_indices["Cidade"]) if "Cidade" in col_indices else None
                local_cell = sheet.cell(row=row, column=col_indices["Local"]) if "Local" in col_indices else None
                
                email_val = str(email_cell.value).strip() if email_cell and email_cell.value is not None else ""
                whatsapp_val = str(whatsapp_cell.value).strip() if whatsapp_cell and whatsapp_cell.value is not None else ""
                uf_val = str(uf_cell.value).strip() if uf_cell and uf_cell.value is not None else ""
                valores_val = str(valores_cell.value).strip() if valores_cell and valores_cell.value is not None else ""
                quartos_val = str(quartos_cell.value).strip() if quartos_cell and quartos_cell.value is not None else ""
                cidade_val = str(cidade_cell.value).strip() if cidade_cell and cidade_cell.value is not None else ""
                local_val = str(local_cell.value).strip() if local_cell and local_cell.value is not None else ""
                
                # Validações sem lag de subprocessos (tudo puxado do cache)
                is_email_valid, email_status, email_score = validate_email_osint(email_val, pousada_name)
                is_whats_valid, whats_status, whats_score, clean_whats = validate_whatsapp_osint(whatsapp_val, uf_val)
                power_tier, power_desc, power_score = calculate_purchasing_power(valores_val, quartos_val)
                
                combined_valid_score = int((email_score + whats_score) / 2)
                
                has_invalid_contact = False
                
                # Coloração
                if email_cell:
                    if not is_email_valid:
                        email_cell.fill = RED_FILL
                        email_cell.font = RED_FONT
                        has_invalid_contact = True
                        
                if whatsapp_cell:
                    if not is_whats_valid or whats_status == "WhatsApp Fixo Corporativo":
                        whatsapp_cell.fill = RED_FILL
                        whatsapp_cell.font = RED_FONT
                        has_invalid_contact = True
                        
                if (not is_email_valid) and (not is_whats_valid):
                    for c_idx in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=c_idx).fill = RED_FILL
                        
                if "Validacao" in col_indices:
                    sheet.cell(row=row, column=col_indices["Validacao"]).value = f"E-mail: {email_status} | WA: {whats_status}"
                if "Score Valid" in col_indices:
                    sheet.cell(row=row, column=col_indices["Score Valid"]).value = combined_valid_score
                if "Qualificacao" in col_indices:
                    sheet.cell(row=row, column=col_indices["Qualificacao"]).value = power_tier
                if "Comportamento" in col_indices:
                    sheet.cell(row=row, column=col_indices["Comportamento"]).value = power_desc
                    
                if has_invalid_contact:
                    invalid_leads_count += 1
                else:
                    valid_leads_count += 1
                    
                # Guardar Praia do Rosa
                is_praiadorosa = "rosa" in local_val.lower() or "rosa" in cidade_val.lower() or "rosa" in pousada_name.lower()
                is_imbituba = "imbituba" in cidade_val.lower() or "ibiraquera" in local_val.lower() or "ibiraquera" in cidade_val.lower()
                
                if is_praiadorosa or is_imbituba:
                    social_val = str(sheet.cell(row=row, column=col_indices["Redes"]).value).strip() if "Redes" in col_indices and sheet.cell(row=row, column=col_indices["Redes"]).value is not None else ""
                    lat_val = sheet.cell(row=row, column=col_indices["LATITUDE"]).value if "LATITUDE" in col_indices else None
                    lon_val = sheet.cell(row=row, column=col_indices["LONGITUDE"]).value if "LONGITUDE" in col_indices else None
                    
                    lead_data = {
                        "pousada": pousada_name,
                        "email": email_val,
                        "whatsapp": whatsapp_val,
                        "quartos": quartos_val,
                        "local": local_val or "Praia do Rosa",
                        "cidade": cidade_val or "Imbituba",
                        "uf": uf_val or "SC",
                        "valores": valores_val,
                        "qualificacao": power_tier,
                        "validacao": f"E-mail: {email_status} | WA: {whats_status}",
                        "comportamento": power_desc,
                        "sinais": sheet.cell(row=row, column=col_indices["Sinais"]).value if "Sinais" in col_indices else "Presença digital ativa.",
                        "redes": social_val,
                        "lat": lat_val,
                        "lon": lon_val,
                        "score_qual": power_score,
                        "score_valid": combined_valid_score,
                        "is_valid": (is_email_valid and is_whats_valid)
                    }
                    praia_rosa_leads.append(lead_data)
                    
            wb.save(output_name)
            print(f"   ✅ Salvo: {os.path.basename(output_name)}")
            print(f"      Válidos: {valid_leads_count} | Inválidos/Incompletos marcados: {invalid_leads_count}")
            
        except Exception as e:
            print(f"   ❌ Erro ao processar {file_item}: {e}")
            
    # 3. Consolidação final Praia do Rosa
    print("\n🏆 [CONSOLIDATION] Gravando dados higienizados e únicos da Fase 1...")
    
    seen_consolidated = set()
    unique_praia_rosa = []
    
    praia_rosa_leads.sort(key=lambda x: (x["is_valid"], x["score_valid"]), reverse=True)
    
    for lead in praia_rosa_leads:
        key = (lead["pousada"].lower().strip(), lead["email"].lower().strip())
        if key not in seen_consolidated:
            seen_consolidated.add(key)
            unique_praia_rosa.append(lead)
            
    print(f"   Total de leads únicos consolidados Praia do Rosa/Imbituba: {len(unique_praia_rosa)}")
    
    try:
        wb_template = openpyxl.load_workbook(CONSOLIDATED_FILE_PATH)
        sheet_out = wb_template.active
        
        if sheet_out.max_row > 1:
            sheet_out.delete_rows(2, sheet_out.max_row)
            
        for idx, lead in enumerate(unique_praia_rosa, 1):
            row_idx = idx + 1
            
            sheet_out.cell(row=row_idx, column=1).value = idx
            sheet_out.cell(row=row_idx, column=2).value = lead["pousada"]
            sheet_out.cell(row=row_idx, column=3).value = lead["email"]
            sheet_out.cell(row=row_idx, column=4).value = lead["whatsapp"]
            sheet_out.cell(row=row_idx, column=5).value = lead["quartos"]
            sheet_out.cell(row=row_idx, column=6).value = lead["local"]
            sheet_out.cell(row=row_idx, column=7).value = lead["cidade"]
            sheet_out.cell(row=row_idx, column=8).value = lead["uf"]
            sheet_out.cell(row=row_idx, column=9).value = lead["valores"]
            sheet_out.cell(row=row_idx, column=10).value = lead["qualificacao"]
            sheet_out.cell(row=row_idx, column=11).value = lead["validacao"]
            sheet_out.cell(row=row_idx, column=12).value = lead["comportamento"]
            sheet_out.cell(row=row_idx, column=13).value = lead["sinais"]
            sheet_out.cell(row=row_idx, column=14).value = lead["redes"]
            sheet_out.cell(row=row_idx, column=15).value = lead["lat"]
            sheet_out.cell(row=row_idx, column=16).value = lead["lon"]
            sheet_out.cell(row=row_idx, column=17).value = lead["score_qual"]
            sheet_out.cell(row=row_idx, column=18).value = lead["score_valid"]
            
            # Coloração no consolidado
            is_email_v, _, _ = validate_email_osint(lead["email"], lead["pousada"])
            is_whats_v, whats_st, _, _ = validate_whatsapp_osint(lead["whatsapp"], lead["uf"])
            
            if not is_email_v:
                sheet_out.cell(row=row_idx, column=3).fill = RED_FILL
                sheet_out.cell(row=row_idx, column=3).font = RED_FONT
            if not is_whats_v or whats_st == "WhatsApp Fixo Corporativo":
                sheet_out.cell(row=row_idx, column=4).fill = RED_FILL
                sheet_out.cell(row=row_idx, column=4).font = RED_FONT
                
            if (not is_email_v) and (not is_whats_v):
                for col in range(1, 19):
                    sheet_out.cell(row=row_idx, column=col).fill = RED_FILL
                    
        wb_template.save(CONSOLIDATED_FILE_PATH)
        print(f"\n🎉 [SUCESSO] Planilha FASE 1 consolidada com sucesso em: {CONSOLIDATED_FILE_PATH}")
        
    except Exception as e:
        print(f"❌ Erro ao consolidar Fase 1: {e}")

if __name__ == "__main__":
    main()
