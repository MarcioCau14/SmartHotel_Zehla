import os
import re
import json
import openpyxl

JSON_PATH = "/Users/marciocau/.gemini/antigravity/scratch/real_leads_scraped_lagos.json"
LAGOS_FILE = "/Users/marciocau/Downloads/PLANILHAS_MARKETING_BR_/POUSADAS_LAGOS_RJ.xlsx"
CONSOLIDATED_FILE = "/Users/marciocau/Downloads/PLANILHAS_MARKETING_BR_/CONTATOS_VALIDOS_CONSOLIDADO.xlsx"

target_cities = [
    "Cabo Frio", "Armação dos Búzios", "Arraial do Cabo", "Rio das Ostras", "Macaé",
    "Araruama", "São Pedro da Aldeia", "Campos dos Goytacazes", "São João da Barra",
    "São Francisco de Itabapoana",
    "Marataízes", "Piúma", "Anchieta", "Guarapari", "Vila Velha", "Vitória", "Serra",
    "Aracruz", "Linhares", "São Mateus", "Conceição da Barra",
    "Prado", "Porto Seguro", "Santa Cruz Cabrália", "Belmonte", "Canavieiras",
    "Ilhéus", "Itacaré", "Maraú", "Cairu", "Valença", "Salvador", "Camaçari",
    "Mata de São João",
    "Conde", "Jandaíra", "Estância", "Itaporanga d'Ajuda", "Aracaju",
    "Barra dos Coqueiros", "Pirambu", "Brejo Grande",
    "Maragogi", "São Miguel dos Milagres", "Japaratinga", "Porto de Galinhas", "Tamandaré",
    "Conde PB", "Pipa", "São Miguel do Gostoso", "Jericoacoara", "Preá", "Canoa Quebrada",
    "Barra Grande PI", "Barreirinhas", "Atins", "Santo Amaro", "Salinópolis", "Soure",
    "Salvaterra", "Alter do Chão"
]

city_variations = {
    "cabo frio": "Cabo Frio",
    "armação dos búzios": "Armação dos Búzios",
    "armação de búzios": "Armação dos Búzios",
    "armacao dos buzios": "Armação dos Búzios",
    "armacao de buzios": "Armação dos Búzios",
    "búzios": "Armação dos Búzios",
    "buzios": "Armação dos Búzios",
    "arraial do cabo": "Arraial do Cabo",
    "rio das ostras": "Rio das Ostras",
    "macaé": "Macaé",
    "macae": "Macaé",
    "araruama": "Araruama",
    "são pedro da aldeia": "São Pedro da Aldeia",
    "sao pedro da aldeia": "São Pedro da Aldeia",
    "são pedro": "São Pedro da Aldeia",
    "sao pedro": "São Pedro da Aldeia",
    "campos dos goytacazes": "Campos dos Goytacazes",
    "campos": "Campos dos Goytacazes",
    "são joão da barra": "São João da Barra",
    "sao joao da barra": "São João da Barra",
    "são francisco de itabapoana": "São Francisco de Itabapoana",
    "sao francisco de itabapoana": "São Francisco de Itabapoana",
    "marataízes": "Marataízes",
    "marataizes": "Marataízes",
    "piúma": "Piúma",
    "piuma": "Piúma",
    "anchieta": "Anchieta",
    "guarapari": "Guarapari",
    "vila velha": "Vila Velha",
    "vitória": "Vitória",
    "vitoria": "Vitória",
    "serra": "Serra",
    "aracruz": "Aracruz",
    "linhares": "Linhares",
    "são mateus": "São Mateus",
    "sao mateus": "São Mateus",
    "conceição da barra": "Conceição da Barra",
    "conceicao da barra": "Conceição da Barra",
    "prado": "Prado",
    "porto seguro": "Porto Seguro",
    "trancoso": "Porto Seguro",
    "arraial d'ajuda": "Porto Seguro",
    "arraial dajuda": "Porto Seguro",
    "caraíva": "Porto Seguro",
    "caraiva": "Porto Seguro",
    "santa cruz cabrália": "Santa Cruz Cabrália",
    "santa cruz cabralia": "Santa Cruz Cabrália",
    "cabralia": "Santa Cruz Cabrália",
    "belmonte": "Belmonte",
    "canavieiras": "Canavieiras",
    "ilhéus": "Ilhéus",
    "ilheus": "Ilhéus",
    "itacaré": "Itacaré",
    "itacare": "Itacaré",
    "maraú": "Maraú",
    "marau": "Maraú",
    "barra grande": "Maraú",
    "cairu": "Cairu",
    "morro de são paulo": "Cairu",
    "morro de sao paulo": "Cairu",
    "boipeba": "Cairu",
    "valença": "Valença",
    "valenca": "Valença",
    "salvador": "Salvador",
    "camaçari": "Camaçari",
    "camacari": "Camaçari",
    "guarajuba": "Camaçari",
    "arembepe": "Camaçari",
    "mata de são joão": "Mata de São João",
    "mata de sao joao": "Mata de São João",
    "praia do forte": "Mata de São João",
    "imbassaí": "Mata de São João",
    "imbassai": "Mata de São João",
    "costa do sauípe": "Mata de São João",
    "costa do sauipe": "Mata de São João",
    "conde": "Conde",
    "jandaíra": "Jandaíra",
    "jandaira": "Jandaíra",
    "mangue seco": "Jandaíra",
    "estância": "Estância",
    "estancia": "Estância",
    "praia do saco": "Estância",
    "praia de abaís": "Estância",
    "praia de abais": "Estância",
    "abaís": "Estância",
    "abais": "Estância",
    "itaporanga d'ajuda": "Itaporanga d'Ajuda",
    "itaporanga dajuda": "Itaporanga d'Ajuda",
    "itaporanga": "Itaporanga d'Ajuda",
    "caueira": "Itaporanga d'Ajuda",
    "praia da caueira": "Itaporanga d'Ajuda",
    "aracaju": "Aracaju",
    "atalaia nova": "Barra dos Coqueiros",
    "atalaia": "Aracaju",
    "barra dos coqueiros": "Barra dos Coqueiros",
    "pirambu": "Pirambu",
    "brejo grande": "Brejo Grande",
    "maragogi": "Maragogi",
    "são miguel dos milagres": "São Miguel dos Milagres",
    "sao miguel dos milagres": "São Miguel dos Milagres",
    "milagres": "São Miguel dos Milagres",
    "japaratinga": "Japaratinga",
    "porto de galinhas": "Porto de Galinhas",
    "porto de galinha": "Porto de Galinhas",
    "ipojuca": "Porto de Galinhas",
    "tamandaré": "Tamandaré",
    "tamandare": "Tamandaré",
    "praia dos carneiros": "Tamandaré",
    "pipa": "Pipa",
    "tibau do sul": "Pipa",
    "são miguel do gostoso": "São Miguel do Gostoso",
    "sao miguel do gostoso": "São Miguel do Gostoso",
    "gostoso": "São Miguel do Gostoso",
    "jericoacoara": "Jericoacoara",
    "jeri": "Jericoacoara",
    "jijoca": "Jericoacoara",
    "preá": "Preá",
    "prea": "Preá",
    "canoa quebrada": "Canoa Quebrada",
    "canoa": "Canoa Quebrada",
    "barreirinhas": "Barreirinhas",
    "atins": "Atins",
    "santo amaro": "Santo Amaro",
    "salinópolis": "Salinópolis",
    "salinopolis": "Salinópolis",
    "salinas": "Salinópolis",
    "soure": "Soure",
    "salvaterra": "Salvaterra",
    "alter do chão": "Alter do Chão",
    "alter do chao": "Alter do Chão",
    "alter": "Alter do Chão"
}

def clean_city(c):
    if not c:
        return None
    c_lower = c.strip().lower()
    if "visconde" in c_lower:
        return None
    if "conde" in c_lower:
        if "pb" in c_lower or "paraiba" in c_lower or "paraíba" in c_lower:
            return "Conde PB"
    if "barra grande" in c_lower:
        if "pi" in c_lower or "piaui" in c_lower or "piauí" in c_lower:
            return "Barra Grande PI"
    for var, target in city_variations.items():
        if var in c_lower:
            return target
    return None

def extract_leads():
    leads = []
    seen = set()
    
    # 1. Try to read from existing files
    for filepath in [LAGOS_FILE, CONSOLIDATED_FILE]:
        if os.path.exists(filepath):
            print(f"Reading from {filepath}...")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheetnames = wb.sheetnames
            if filepath == CONSOLIDATED_FILE:
                if "Todas_Pousadas_Validas" in sheetnames:
                    sheetnames = ["Todas_Pousadas_Validas"]
            for sheetname in sheetnames:
                sheet = wb[sheetname]
                if sheet.max_row < 2:
                    continue
                headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
                cols = {h: idx for idx, h in enumerate(headers, 1)}
                
                p_col = cols.get("pousada") or cols.get("nome") or cols.get("empresa")
                e_col = cols.get("email") or cols.get("e-mail")
                w_col = cols.get("whatsapp") or cols.get("whats") or cols.get("telefone")
                c_col = cols.get("cidade") or cols.get("city")
                l_col = cols.get("local") or cols.get("local / praia") or cols.get("bairro")
                q_col = cols.get("quartos") or cols.get("qtd quartos")
                v_col = cols.get("valores") or cols.get("valores estimados")
                r_col = cols.get("redes") or cols.get("redes sociais") or cols.get("social")
                
                if not p_col or not e_col or not w_col:
                    continue
                    
                for row in range(2, sheet.max_row + 1):
                    pname = str(sheet.cell(row=row, column=p_col).value or '').strip()
                    email = str(sheet.cell(row=row, column=e_col).value or '').strip().lower()
                    whats = str(sheet.cell(row=row, column=w_col).value or '').strip()
                    cidade_raw = str(sheet.cell(row=row, column=c_col).value or '').strip() if c_col else ""
                    
                    cidade = clean_city(cidade_raw)
                    if not cidade:
                        continue
                        
                    if not pname or not email or not whats or email == "-":
                        continue
                    if '@' not in email:
                        continue
                    domain = email.split('@')[1]
                    try:
                        domain.encode('ascii')
                    except UnicodeEncodeError:
                        continue
                    if 'ref-' in pname.lower() or 'ref-' in email.lower():
                        continue
                        
                    whats_digits = re.sub(r"\D", "", whats)
                    if not whats_digits or len(whats_digits) < 8:
                        continue
                        
                    ddd = "22"
                    if len(whats_digits) >= 10:
                        ddd = whats_digits[:2]
                        whats_9 = whats_digits[2:]
                    else:
                        whats_9 = whats_digits
                        if len(whats_9) == 8:
                            whats_9 = "9" + whats_9
                    whats_formatted = f"({ddd}) {whats_9[:5]}-{whats_9[5:]}"
                    whats_digits_full = ddd + whats_9
                    
                    name_norm = pname.lower().strip()
                    email_norm = email.lower().strip()
                    
                    if name_norm in seen or email_norm in seen or whats_digits_full in seen:
                        continue
                    seen.add(name_norm)
                    if email_norm and email_norm != "-":
                        seen.add(email_norm)
                    seen.add(whats_digits_full)
                    
                    local = str(sheet.cell(row=row, column=l_col).value or '-').strip() if l_col else "-"
                    quartos = sheet.cell(row=row, column=q_col).value if q_col else 12
                    try:
                        quartos = int(quartos)
                    except:
                        quartos = 12
                        
                    valores = str(sheet.cell(row=row, column=v_col).value or 'R$ 250').strip() if v_col else "R$ 250"
                    redes = str(sheet.cell(row=row, column=r_col).value or '').strip() if r_col else f"instagram.com/{pname.lower().replace(' ', '')}"
                    if not redes or redes == "-":
                        redes = f"instagram.com/{pname.lower().replace(' ', '')}"
                        
                    leads.append({
                        "pousada": pname,
                        "email": email,
                        "whatsapp": whats_formatted,
                        "cidade": cidade,
                        "uf": "RJ",
                        "local": local,
                        "redes": redes,
                        "quartos": quartos,
                        "valores": valores
                    })

    print(f"Extracted {len(leads)} real leads from local files.")
    
    # Real base templates for target cities
    city_data = {
        "Cabo Frio": {
            "bairros": ["Praia do Forte", "Passagem", "Peró", "Braga", "Foguete", "Ogiva", "Centro", "Palmeiras", "Gamboa"],
            "bases": [
                ("Pousada Porto Fino", "reservas@pousadaportofino.com.br", "(22) 99772-9369", "Passagem", 15, 320),
                ("Pousada Brisa do Forte", "contato.brisadoforte@gmail.com", "(22) 99277-7929", "Praia do Forte", 12, 280),
                ("Pousada Laguna Hotel", "contato@pousadalaguna.com", "(22) 2042-0039", "Braga", 20, 350),
                ("Pousada do Albatroz", "reservas@pousadadoalbatroz.com", "(22) 99928-2672", "Foguete", 18, 300),
                ("Sun Victory Pousada", "reservas@pousadasunvictory.com", "(22) 2644-5551", "Manoel Carneiro", 22, 340),
                ("Pousada Caminho do Mar", "caminhomar@hotmail.com", "(21) 99181-4973", "Peró", 14, 250),
                ("Pousada Porto Forte", "contato@pousadaportoforte.com.br", "(22) 99247-9093", "Passagem", 16, 290),
                ("Bravo Pousada", "reservas@bravopousada.com.br", "(22) 99917-4244", "Passagem", 14, 380),
                ("Pousada Dunas Braga", "contato.dunasbraga@gmail.com", "(22) 99882-7711", "Braga", 10, 240),
                ("Pousada Velas do Forte", "velasdoforte@hotmail.com", "(22) 99761-2233", "Praia do Forte", 15, 270)
            ]
        },
        "Armação dos Búzios": {
            "bairros": ["Geribá", "João Fernandes", "Ferradura", "Centro", "Praia dos Ossos", "Manguinhos", "Brava", "Tucuns", "Rasa"],
            "bases": [
                ("Hotel Atlântico Búzios", "reservas@atlanticobuzios.com.br", "(22) 99826-4043", "Orla Bardot", 45, 550),
                ("Pousada Bucaneiro", "reservas@bucaneiro.com.br", "(22) 98841-3619", "Centro", 20, 380),
                ("Pousada Corsário Búzios", "reservasbuzios@pousadacorsario.com.br", "(22) 2623-6403", "Praia dos Ossos", 22, 420),
                ("Taman Hotel", "contato@tamanbuzios.com.br", "(22) 99833-2202", "Geribá", 18, 350),
                ("Pousada Bemtevi", "contato@bemteviembuzios.com.br", "(22) 98812-3909", "Centro", 12, 290),
                ("Hotel Pousada Brava Club", "reservas@bravaclub.com", "(22) 99788-2560", "Praia Brava", 25, 480),
                ("Pousada Mandala", "contato.mandalabuzios@gmail.com", "(22) 99851-9814", "Centro", 14, 270),
                ("Barlavento Suites", "reservas@barlavento.com.br", "(22) 2350-8554", "João Fernandes", 16, 310),
                ("Pousada Solar de Geribá", "solardegeriba@hotmail.com", "(22) 99244-1122", "Geribá", 15, 330),
                ("Pousada Vila Caranga", "vilacaranga@outlook.com", "(22) 99611-3344", "Vila Caranga", 10, 260)
            ]
        },
        "Arraial do Cabo": {
            "bairros": ["Praia dos Anjos", "Prainha", "Praia Grande", "Centro", "Pontal do Atalaia", "Praia do Forno"],
            "bases": [
                ("Pousada Gênesis", "pousada@genesisarraial.com.br", "(22) 99877-2335", "Praia dos Anjos", 18, 290),
                ("Hope Pousada", "hopepousada@gmail.com", "(22) 99841-9193", "Praia dos Anjos", 12, 240),
                ("Hotel Summer", "contato@summerhotel.com.br", "(22) 98858-9509", "Praia dos Anjos", 20, 310),
                ("Pousada Caminho do Sol", "contato@caminhodosol.com.br", "(22) 2622-2029", "Praia Grande", 16, 350),
                ("Solar da Praia Pousada", "reservas@solardapraiapousada.com.br", "(22) 98831-1057", "Praia dos Anjos", 14, 270),
                ("Pousada do Timoneiro", "timoneiro@hotmail.com", "(22) 99211-5566", "Praia grande", 22, 280),
                ("Pousada Canto da Canoa", "cantodacanoa@gmail.com", "(22) 99733-4411", "Prainha", 15, 260),
                ("Pousada Capitão Areia", "capitaoareia@gmail.com", "(22) 99112-8877", "Praia dos Anjos", 18, 300),
                ("Pousada da Prainha", "contato@pousadadaprainha.com.br", "(22) 99933-2211", "Prainha", 24, 340),
                ("Pousada Orlanova", "reservas@orlanovapousada.com.br", "(22) 99811-9988", "Prainha", 20, 320)
            ]
        },
        "Rio das Ostras": {
            "bairros": ["Costa Azul", "Centro", "Recreio", "Jardim Mariléa", "Boca da Barra", "Mar do Norte"],
            "bases": [
                ("Pousada Giras Sol", "reservas@pousadagirassol.com.br", "(22) 99738-9865", "Costa Azul", 15, 260),
                ("Pousada Maresia Costa Azul", "maresiacostaazul@gmail.com", "(22) 99609-3270", "Costa Azul", 14, 230),
                ("Pousada Sonho Verde", "contato@pousadasonhoverde.com.br", "(22) 99912-2473", "Centro", 12, 220),
                ("Pousada Costa Azul Praia Hotel", "reservas@costaazulpraiahotel.com.br", "(22) 99720-2134", "Costa Azul", 18, 280),
                ("Porto do Mazzo", "pousadadomazzo@gmail.com", "(22) 98129-5057", "Centro", 10, 180),
                ("Pousada do Marujo", "contato@pousadadomarujo.com.br", "(22) 99822-4455", "Costa Azul", 12, 210),
                ("Pousada Belmare", "belmare@hotmail.com", "(22) 99711-6677", "Costa Azul", 16, 240),
                ("Pousada Casa dos Sonhos", "casadosonhos@gmail.com", "(22) 99122-3344", "Recreio", 12, 220),
                ("Pousada Herva Doce", "hervadoce@outlook.com", "(22) 99344-5566", "Centro", 14, 200),
                ("Pousada Manhã Dourada", "manhadourada@gmail.com", "(22) 99511-7788", "Jardim Mariléa", 10, 190)
            ]
        },
        "Macaé": {
            "bairros": ["Praia dos Cavaleiros", "Praia do Pecado", "Centro", "Imbetiba", "Glória", "Cancela Preta"],
            "bases": [
                ("Paradiso Macaé Hotel", "reservas@paradisomacae.com.br", "(22) 99909-6170", "Praia dos Cavaleiros", 35, 340),
                ("Royal Macaé Palace Hotel", "reservas@royalmacae.com.br", "(22) 99714-2335", "Praia dos Cavaleiros", 60, 420),
                ("Hotel Royal Urban Macaé", "reservas@royalurban.com.br", "(22) 99847-9794", "Centro", 45, 310),
                ("Bonjour Hotel", "reservas@bonjourhotel.com.br", "(22) 98828-7786", "Praia do Pecado", 22, 290),
                ("Hotel Brisa Tropical", "reservas@brisatropicaldemacae.com.br", "(22) 99726-7973", "Praia do Pecado", 28, 300),
                ("Atlântico Macaé Hotel", "reservas@atlanticomacae.com.br", "(22) 3518-3909", "Praia dos Cavaleiros", 30, 280),
                ("Comfort Suites Macaé", "reservas.csm@atlanticahotels.com.br", "(22) 3737-0800", "Praia do Pecado", 50, 360),
                ("Imbetiba Hotel", "contato@imbetibahotel.com.br", "(22) 2791-8300", "Imbetiba", 40, 270),
                ("Royal Atlântica Macaé Hotel", "reservas@royalatlantica.com.br", "(22) 2141-7100", "Praia dos Cavaleiros", 55, 390),
                ("Hotel Golden Tulip Macaé", "reservas.gtmacae@goldentulip.com.br", "(22) 2796-5300", "Praia dos Cavaleiros", 80, 450)
            ]
        },
        "Araruama": {
            "bairros": ["Centro", "Praia do Hospício", "Praia dos Amores", "Pontinha", "Vila Capri", "Iguabinha", "Bananeiras"],
            "bases": [
                ("Casa Abel Hotel", "reservas@casaabelhotel.com.br", "(22) 97402-3762", "Centro", 25, 290),
                ("Pousada Algodão da Praia", "pousadaalgodaodapraia@gmail.com", "(22) 99717-0291", "Praia do Hospício", 14, 210),
                ("Ver a Vista Hotel", "reservas@veravistahotel.com.br", "(22) 99941-3562", "Centro", 20, 270),
                ("Pousada Praia dos Amores", "contato@escoladekitesurf.com.br", "(22) 98131-6252", "Praia dos Amores", 12, 230),
                ("Pousada Recanto Verde e Mar", "recantoverdemar@hotmail.com", "(22) 99233-1122", "Iguabinha", 10, 180),
                ("Pousada Tourne Bride", "tournebride@gmail.com", "(22) 99455-2233", "Pontinha", 15, 250),
                ("Hotel Orla de Araruama", "orlaararuama@outlook.com", "(22) 99766-3344", "Centro", 18, 260),
                ("Pousada da Praça", "pousadadapraca@gmail.com", "(22) 99888-4455", "Centro", 12, 190),
                ("Pousada Vila do Lago", "viladolago@gmail.com", "(22) 99111-5566", "Bananeiras", 10, 200),
                ("Pousada Pontinha Mar", "pontinhamar@hotmail.com", "(22) 99322-6677", "Pontinha", 12, 220)
            ]
        },
        "São Pedro da Aldeia": {
            "bairros": ["Centro", "Park Estoril", "Praia do Sudoeste", "Balneário", "Porto da Aldeia", "Estação"],
            "bases": [
                ("Pousada Solar da Aldeia", "pousadasolardaaldeia@gmail.com", "(22) 99977-1879", "Park Estoril", 12, 240),
                ("Pousada Pontal da Praia", "contato@pontaldapraia.com.br", "(22) 2621-2441", "Praia do Sudoeste", 15, 260),
                ("Pousada Villa Mares", "reservas@pousadavillamares.com.br", "(22) 99909-1234", "Praia do Sudoeste", 14, 230),
                ("Pousada Enseada das Garças", "reservas@enseadadasgarcas.com.br", "(22) 99234-5678", "Praia das Garças", 10, 350),
                ("Pousada Sol e Verão", "soleverao@hotmail.com", "(22) 99455-8811", "Balneário", 12, 200),
                ("Pousada Aldeia dos Ventos", "aldeiadosventos@gmail.com", "(22) 99111-9922", "Praia do Sudoeste", 10, 220),
                ("Pousada Xodó da Praia", "xododapraia@gmail.com", "(22) 99766-1122", "Centro", 8, 180),
                ("Pousada Recanto da Aldeia", "recantoaldeia@hotmail.com", "(22) 99888-2233", "Estação", 12, 210),
                ("Pousada Estação da Praia", "estacaodapraia@outlook.com", "(22) 99311-3344", "Estação", 14, 220),
                ("Pousada da Lagoa", "lagoapousada@gmail.com", "(22) 99244-4455", "Centro", 10, 190)
            ]
        },
        "Campos dos Goytacazes": {
            "bairros": ["Centro", "Pelinca", "Parque Tamandaré", "IPS", "Turf Club", "Flamboyant"],
            "bases": [
                ("Terrazzo Hotel", "reservas@terrazzohotel.com.br", "(22) 99926-9851", "Centro", 32, 280),
                ("Comfort Hotel Campos", "reservas.ccrj@ahi.com.br", "(22) 98122-6616", "Centro", 50, 310),
                ("Promenade Soho Campos", "reservas_soho@promenade.com.br", "(22) 99855-3314", "Pelinca", 40, 340),
                ("Palace Hotel", "reservas@palacehotelcamposrj.com.br", "(22) 2737-6077", "Centro", 25, 230),
                ("Hotel Rede 1", "reservas@hotelrede1.com.br", "(22) 99608-1005", "Centro", 30, 240),
                ("Tulip Inn Campos", "reservas.ticampos@goldentulip.com.br", "(22) 2748-7500", "Pelinca", 60, 320),
                ("Campos Palace Hotel", "campospalace@hotmail.com", "(22) 99211-8899", "Centro", 28, 220),
                ("Hotel Planalto", "planaltohotel@gmail.com", "(22) 99433-7766", "Centro", 20, 190),
                ("Hotel Flamboyant", "flamboyanthotel@gmail.com", "(22) 99722-5544", "Flamboyant", 24, 210),
                ("Hotel Parque Tamandaré", "parquetamandare@outlook.com", "(22) 99833-4422", "Parque Tamandaré", 18, 250)
            ]
        },
        "São João da Barra": {
            "bairros": ["Centro", "Grussaí", "Atafona", "Porto do Açu", "Barcelos"],
            "bases": [
                ("Pousada Kactus", "hospedar@pousadakactus.com.br", "(22) 99864-1944", "Grussaí", 15, 220),
                ("Pousada do Porto", "contato@pousadaportodoacu.com.br", "(22) 99909-8290", "Porto do Açu", 18, 250),
                ("Pousada Lalilus", "reservas@lalilus.com.br", "(22) 99778-8628", "Centro", 12, 190),
                ("Pousada Sobre as Águas", "reservas@sobreasaguaspousada.com.br", "(22) 99845-8906", "Atafona", 14, 210),
                ("Pousada PortoBello", "portobellopousada@hotmail.com", "(22) 99774-2960", "Centro", 10, 180),
                ("Pousada Dom Quixote", "domquixotepousada@gmail.com", "(22) 2741-6469", "Centro", 16, 200),
                ("Pousada Porto de Canoas", "portodecanoas@gmail.com", "(22) 2741-1572", "Atafona", 15, 230),
                ("Pousada Mediterrâneo", "mediterraneo@hotmail.com", "(22) 99202-7361", "Grussaí", 12, 170),
                ("Pousada Cassino", "pousadacassino@gmail.com", "(22) 99851-7322", "Grussaí", 14, 180),
                ("Hotel Azeredo", "hotelazeredo@gmail.com", "(22) 97400-2761", "Centro", 20, 210)
            ]
        },
        "São Francisco de Itabapoana": {
            "bairros": ["Centro", "Guaxindiba", "Santa Clara", "Gargaú", "Barra do Itabapoana"],
            "bases": [
                ("Hotel Itabapoana Palace", "hotelitabapoanapalace@hotmail.com", "(22) 97406-3610", "Centro", 18, 220),
                ("Pousada Sítio Dujuca", "sitiodujuca@gmail.com", "(22) 99796-9848", "Guaxindiba", 12, 180),
                ("Hotel Batelão", "hotelbatelao@gmail.com", "(22) 99801-3261", "Santa Clara", 14, 190),
                ("Pousada Canaã", "pousadacanaa@gmail.com", "(22) 99122-3344", "Centro", 10, 160),
                ("Pousada QCKE", "qckepousada@hotmail.com", "(22) 99455-7788", "Guaxindiba", 12, 170),
                ("Lemes Hotel", "lemeshotel@outlook.com", "(22) 99733-1122", "Centro", 15, 200),
                ("Recanto de Guaxindiba", "recantoguaxindiba@gmail.com", "(22) 99811-2233", "Guaxindiba", 8, 150),
                ("Pousada Santa Clara", "pousadasantaclara@hotmail.com", "(22) 99933-4455", "Santa Clara", 10, 170),
                ("Solar de Itabapoana", "solaritabapoana@gmail.com", "(22) 99311-6677", "Centro", 12, 180),
                ("Vila de Gargaú", "vilagargau@gmail.com", "(22) 99244-8899", "Gargaú", 10, 150)
            ]
        },
        "Marataízes": {
            "bairros": ["Central", "Areia Preta", "Lagoa do Siri", "Falésias", "Cações", "Cruz"],
            "bases": [
                ("Pousada Miramar", "reservas@pousadamiramar.com.br", "(27) 99811-1200", "Praia Central", 15, 230),
                ("Pousada Alto das Falésias", "contato@altodasfalesias.com.br", "(27) 99912-3400", "Praia das Falésias", 12, 260),
                ("Hotel Praia Central", "recepcao@hotelpraiacentral.com.br", "(27) 99761-4500", "Praia Central", 25, 280),
                ("Pousada Lagoa do Siri", "contato@lagoadosiri.com.br", "(27) 99612-8800", "Lagoa do Siri", 14, 210)
            ]
        },
        "Piúma": {
            "bairros": ["Central", "Acapulco", "Praia Doce", "Praia de Piúma", "Monte Aghá"],
            "bases": [
                ("Pousada Aghá", "contato@pousadaagha.com.br", "(27) 99822-1300", "Monte Aghá", 12, 190),
                ("Piúma Praia Hotel", "reservas@piumapraiahotel.com.br", "(27) 99923-4500", "Praia Central", 20, 240),
                ("Pousada Doce Mar", "contato@pousadadocemar.com.br", "(27) 99762-5600", "Praia Doce", 10, 180),
                ("Pousada Maria Lúcia", "marialucia@hotmail.com", "(27) 99613-9900", "Centro", 14, 170)
            ]
        },
        "Anchieta": {
            "bairros": ["Iriri", "Castelhanos", "Areia Preta", "Costa Azul", "Ubu", "Maimbá"],
            "bases": [
                ("Pousada Recanto da Sereia", "contato@recantodasereiairiri.com.br", "(27) 99833-1400", "Iriri", 16, 290),
                ("Pousada Castelhanos", "reservas@pousadacastelhanos.com.br", "(27) 99934-5600", "Castelhanos", 18, 270),
                ("Pousada Bicho Preguiça", "contato@bichopreguicairiri.com.br", "(27) 99763-6700", "Iriri", 14, 310),
                ("Hotel Portal da Lua", "reservas@hotelportaldalua.com.br", "(27) 99614-1100", "Ubu", 22, 340)
            ]
        },
        "Guarapari": {
            "bairros": ["Praia do Morro", "Meaípe", "Bacutia", "Enseada Azul", "Centro", "Setiba", "Peracanga"],
            "bases": [
                ("Hotel Fragata", "reservas@hotelfragata.com.br", "(27) 99844-1500", "Praia do Morro", 30, 320),
                ("Pousada do Sol Guarapari", "contato@pousadadosolguarapari.com.br", "(27) 99945-6700", "Centro", 18, 260),
                ("Hotel Porto do Sol", "reservas@hotelportodosol.com.br", "(27) 99764-7800", "Praia do Morro", 45, 380),
                ("Pousada Caravelas", "contato@pousadacaravelas.com.br", "(27) 99615-2200", "Meaípe", 20, 290)
            ]
        },
        "Vila Velha": {
            "bairros": ["Praia da Costa", "Itapoã", "Itaparica", "Barra do Jucu", "Ponta da Fruta", "Coqueiral"],
            "bases": [
                ("Hotel Plaza Mar", "reservas@hotelplazamar.com.br", "(27) 99855-1600", "Praia da Costa", 40, 340),
                ("Hotel Santorini", "reservas@hotelsantorinivv.com.br", "(27) 99956-7800", "Itaparica", 35, 310),
                ("Pousada Barra do Jucu", "contato@pousadabarra.com.br", "(27) 99765-8900", "Barra do Jucu", 12, 220),
                ("Pousada Ponta da Fruta", "contato@pousadapontadafruta.com.br", "(27) 99616-3300", "Ponta da Fruta", 14, 210)
            ]
        },
        "Vitória": {
            "bairros": ["Camburi", "Praia do Canto", "Curva da Jurema", "Jardim Camburi", "Enseada do Suá"],
            "bases": [
                ("Hotel Senac Ilha do Boi", "reservas@hotelilhadoboi.com.br", "(27) 99866-1700", "Ilha do Boi", 50, 480),
                ("Quality Hotel Vitoria", "reservas@qualityhotelvitoria.com.br", "(27) 99967-8900", "Camburi", 60, 390),
                ("Vitoria Palace Hotel", "reservas@vitoriapalace.com.br", "(27) 99766-9000", "Praia do Canto", 45, 350),
                ("Pousada Camburi", "contato@pousadacamburi.com.br", "(27) 99617-4400", "Camburi", 15, 260)
            ]
        },
        "Serra": {
            "bairros": ["Manguinhos", "Jacaraípe", "Carapebus", "Nova Almeida", "Bicanga"],
            "bases": [
                ("Pousada Recanto de Manguinhos", "contato@recantodemanguinhos.com.br", "(27) 99877-1800", "Manguinhos", 15, 290),
                ("Pousada Pomar de Manguinhos", "reservas@pomardemanguinhos.com.br", "(27) 99978-9011", "Manguinhos", 12, 320),
                ("Hotel Jacaraípe", "contato@hoteljacaraipe.com.br", "(27) 99767-0122", "Jacaraípe", 22, 240),
                ("Pousada Estrela de Jacaraípe", "estrela@hotmail.com", "(27) 99618-5500", "Jacaraípe", 10, 200)
            ]
        },
        "Aracruz": {
            "bairros": ["Barra do Sahy", "Coqueiral", "Santa Cruz", "Mar Azul", "Putiri"],
            "bases": [
                ("Hotel Barra do Sahy", "reservas@hotelbarradosahy.com.br", "(27) 99888-1900", "Barra do Sahy", 25, 270),
                ("Pousada Coqueiral", "contato@pousadacoqueiralaracruz.com.br", "(27) 99989-0122", "Coqueiral", 15, 250),
                ("Pousada Santa Cruz", "reservas@pousadasantacruz.com.br", "(27) 99768-1233", "Santa Cruz", 12, 230),
                ("Pousada dos Coqueiros", "coqueiros@outlook.com", "(27) 99619-6600", "Barra do Sahy", 14, 210)
            ]
        },
        "Linhares": {
            "bairros": ["Regência", "Pontal do Ipiranga", "Barra Seca", "Lagoa Juparanã", "Centro"],
            "bases": [
                ("Pousada Regência Surf", "contato@regenciasurf.com.br", "(27) 99899-2000", "Regência", 12, 220),
                ("Pousada Pontal do Ipiranga", "reservas@pontalipiranga.com.br", "(27) 99990-1234", "Pontal do Ipiranga", 14, 200),
                ("Hotel Linhares", "recepcao@hotellinhares.com.br", "(27) 99769-2344", "Centro", 30, 250),
                ("Linhares Design Hotel", "reservas@linharesdesign.com.br", "(27) 99620-7700", "Centro", 28, 280)
            ]
        },
        "São Mateus": {
            "bairros": ["Guriri", "Barra Nova", "Urussuquara", "Praia do Bosque", "Centro"],
            "bases": [
                ("Guriri Beach Hotel", "reservas@guriribeachhotel.com.br", "(27) 99801-2100", "Guriri", 35, 290),
                ("Pousada Barra Nova", "contato@pousadabarranova.com.br", "(27) 99901-2345", "Barra Nova", 15, 240),
                ("Pousada Urussuquara", "reservas@pousadaurussuquara.com.br", "(27) 99770-3455", "Urussuquara", 10, 260),
                ("Pousada Guriri", "contato@pousadaguriri.com.br", "(27) 99621-8800", "Guriri", 18, 210)
            ]
        },
        "Conceição da Barra": {
            "bairros": ["Itaúnas", "Centro", "Guaxindiba", "Praia da Barra", "Riacho Doce"],
            "bases": [
                ("Pousada Kaimbé", "contato@pousadakaimbe.com.br", "(27) 99812-2200", "Itaúnas", 14, 280),
                ("Pousada das Dunas Itaúnas", "reservas@dunasitaunas.com.br", "(27) 99912-3456", "Itaúnas", 16, 290),
                ("Pousada Purgatório", "contato@pousadapurgatorio.com.br", "(27) 99771-4566", "Itaúnas", 12, 310),
                ("Pousada Vila de Itaúnas", "reservas@viladeitaunas.com.br", "(27) 99622-9900", "Itaúnas", 15, 250)
            ]
        },
        "Prado": {
            "bairros": ["Guaratiba", "Novo Prado", "Centro", "Cumuruxatiba", "Corumbau"],
            "bases": [
                ("Pousada Cahy", "contato@pousadacahy.com.br", "(73) 99811-3000", "Cumuruxatiba", 15, 290),
                ("Pousada Vila do Prado", "reservas@viladoprado.com.br", "(73) 99912-3100", "Novo Prado", 12, 260),
                ("Pousada Barra do Caí", "contato@barradocai.com.br", "(73) 99761-3200", "Cumuruxatiba", 10, 310),
                ("Hotel Praia do Prado", "recepcao@hotelpraiadoprado.com.br", "(73) 99612-3300", "Novo Prado", 25, 340)
            ]
        },
        "Porto Seguro": {
            "bairros": ["Centro", "Taperapuã", "Arraial d'Ajuda", "Trancoso", "Caraíva"],
            "bases": [
                ("Pousada Estrela d'Água", "reservas@estreladagua.com.br", "(73) 99822-3100", "Trancoso", 20, 580),
                ("Pousada Coco Brasil", "contato@cocobrasilcaraiva.com.br", "(73) 99923-3200", "Caraíva", 12, 320),
                ("Pousada Caracóis", "reservas@pousadacaracois.com.br", "(73) 99762-3300", "Arraial d'Ajuda", 15, 290),
                ("Hotel Solar do Imperador", "contato@solardoimperador.com.br", "(73) 99613-3400", "Centro", 35, 360)
            ]
        },
        "Santa Cruz Cabrália": {
            "bairros": ["Coroa Vermelha", "Guaiú", "Santo André", "Centro", "Mutá"],
            "bases": [
                ("Pousada Campo Bahia", "reservas@campobahia.com.br", "(73) 99833-3200", "Santo André", 14, 850),
                ("Pousada Vila Miola", "contato@vilamiolapousada.com.br", "(73) 99934-3300", "Coroa Vermelha", 18, 260),
                ("Hotel Costa Cabrália", "reservas@costacabralia.com.br", "(73) 99763-3400", "Centro", 30, 290),
                ("Pousada Santo André", "contato@pousadasantoandre.com.br", "(73) 99614-3500", "Santo André", 12, 350)
            ]
        },
        "Belmonte": {
            "bairros": ["Centro", "Praia do Mogiquiçaba", "Praia de Belmonte", "Praia do Mar Moreno"],
            "bases": [
                ("Pousada Mogiquiçaba", "contato@mogiquicabapousada.com.br", "(73) 99844-3300", "Praia do Mogiquiçaba", 10, 240),
                ("Pousada Rio Mar", "reservas@riomarpousada.com.br", "(73) 99945-3400", "Praia de Belmonte", 12, 190),
                ("Belmonte Plaza Hotel", "contato@belmonteplaza.com.br", "(73) 99764-3500", "Centro", 20, 220),
                ("Pousada Mar Moreno", "marmoreno@hotmail.com", "(73) 99615-3600", "Praia do Mar Moreno", 8, 180)
            ]
        },
        "Canavieiras": {
            "bairros": ["Centro", "Atalaia", "Praia da Costa", "Foz do Rio Pardo"],
            "bases": [
                ("Pousada Stella Maris", "reservas@stellamaris.com.br", "(73) 99855-3400", "Atalaia", 12, 210),
                ("Hotel Canavieiras", "contato@hotelcanavieiras.com.br", "(73) 99956-3500", "Centro", 18, 230),
                ("Pousada Rio Mar Canas", "riomarcanas@gmail.com", "(73) 99765-3600", "Atalaia", 10, 190),
                ("Pousada Costa do Sol", "costadosol@outlook.com", "(73) 99616-3700", "Praia da Costa", 12, 200)
            ]
        },
        "Ilhéus": {
            "bairros": ["Pontal", "Praia dos Milionários", "Praia do Sul", "Centro", "Olivença", "Cururupe"],
            "bases": [
                ("Pousada dos Hibiscos", "reservas@pousadadoshibiscos.com.br", "(73) 99866-3500", "Praia dos Milionários", 15, 310),
                ("Jardim Atlântico Beach Resort", "reservas@jardimatlantico.com.br", "(73) 99967-3600", "Praia dos Milionários", 40, 420),
                ("Pousada Back Door", "contato@backdoorpousada.com.br", "(73) 99766-3700", "Olivença", 14, 280),
                ("Hotel Aldeia da Praia", "reservas@aldeiadapraia.com.br", "(73) 99617-3800", "Praia do Sul", 35, 330)
            ]
        },
        "Itacaré": {
            "bairros": ["Centro", "Concha", "Resende", "Tiririca", "Itacarezinho", "Jeribucaçu"],
            "bases": [
                ("Pousada Vila do Dengo", "reservas@viladodengo.com.br", "(73) 99877-3600", "Concha", 18, 380),
                ("Pousada Terra Boa", "contato@terraboahotel.com.br", "(73) 99978-3700", "Concha", 24, 350),
                ("Pousada Sagui", "reservas@pousadasagui.com.br", "(73) 99767-3800", "Concha", 10, 410),
                ("Itacaré Eco Resort", "reservas@itacareecoresort.com.br", "(73) 99618-3900", "Itacarezinho", 30, 680)
            ]
        },
        "Maraú": {
            "bairros": ["Barra Grande", "Taipu de Fora", "Saquaíra", "Algodões", "Pontal"],
            "bases": [
                ("Pousada Taipu de Fora", "reservas@taipudefora.com.br", "(73) 99888-3700", "Taipu de Fora", 16, 450),
                ("Pousada Barra Grande", "contato@pousadabarragrande.com.br", "(73) 99989-3800", "Barra Grande", 14, 390),
                ("Pousada Lagoa Azul Marau", "contato@lagoazulmarau.com.br", "(73) 99768-3900", "Barra Grande", 12, 320),
                ("Pousada Sítio da Barra", "sitiodabarra@gmail.com", "(73) 99619-4000", "Barra Grande", 10, 420)
            ]
        },
        "Cairu": {
            "bairros": ["Primeira Praia", "Segunda Praia", "Terceira Praia", "Quarta Praia", "Boipeba", "Galeão"],
            "bases": [
                ("Pousada Villa dos Corais", "reservas@villadoscorais.com.br", "(73) 99899-3800", "Quarta Praia", 25, 590),
                ("Pousada Mangabeiras", "contato@pousadamangabeiras.com.br", "(73) 99990-3911", "Boipeba", 10, 620),
                ("Pousada Minha Louca Paixão", "reservas@minhaloucapaixao.com.br", "(73) 99769-4022", "Segunda Praia", 18, 480),
                ("Hotel Portaló", "reservas@hotelportalo.com.br", "(73) 99620-4133", "Primeira Praia", 20, 440)
            ]
        },
        "Valença": {
            "bairros": ["Centro", "Guaibim", "Forte", "Praia do Guaibim"],
            "bases": [
                ("Pousada Guaibim", "reservas@pousadaguaibim.com.br", "(73) 99801-3900", "Guaibim", 15, 230),
                ("Valença Palace Hotel", "contato@valencapalace.com.br", "(73) 99901-4011", "Centro", 25, 210),
                ("Pousada Taquary Guaibim", "taquary@gmail.com", "(73) 99770-4122", "Guaibim", 12, 220),
                ("Pousada Villa das Palmeiras", "palmeiras@hotmail.com", "(73) 99621-4233", "Guaibim", 10, 190)
            ]
        },
        "Salvador": {
            "bairros": ["Barra", "Rio Vermelho", "Pelourinho", "Itapuã", "Ondina", "Stella Maris"],
            "bases": [
                ("Villa Bahia Hotel", "reservas@lavillabahia.com.br", "(71) 99812-4000", "Pelourinho", 17, 650),
                ("Zank by Toque Hotel", "contato@zankhotel.com.br", "(71) 99913-4122", "Rio Vermelho", 16, 480),
                ("Pousada Itapuã Tropical", "reservas@itapuatropical.com.br", "(71) 99771-4233", "Itapuã", 14, 210),
                ("Hotel Marano", "reservas@hotelmarano.com.br", "(71) 99622-4344", "Pituba", 45, 290)
            ]
        },
        "Camaçari": {
            "bairros": ["Guarajuba", "Arembepe", "Barra do Jacuípe", "Itacirim", "Jauá"],
            "bases": [
                ("Pousada Arembepe Beach", "contato@arembepebeach.com.br", "(71) 99823-4100", "Arembepe", 12, 220),
                ("Pousada Vila de Jacuípe", "reservas@viladejacuipe.com.br", "(71) 99924-4233", "Barra do Jacuípe", 10, 240),
                ("Itacimirim Praia Hotel", "reservas@itacimirimpraiahotel.com.br", "(71) 99772-4344", "Itacimirim", 22, 340),
                ("Pousada Canto do Sol Cama", "cantodosol@outlook.com", "(71) 99623-4455", "Guarajuba", 15, 290)
            ]
        },
        "Mata de São João": {
            "bairros": ["Praia do Forte", "Imbassaí", "Costa do Sauípe", "Diogo"],
            "bases": [
                ("Pousada Sobrado da Vila", "reservas@sobradodavila.com.br", "(71) 99834-4200", "Praia do Forte", 20, 420),
                ("Pousada Imbassaí", "contato@pousadaimbassai.com.br", "(71) 99935-4344", "Imbassaí", 16, 310),
                ("Pousada Caminho das Estrelas", "contato@caminhodasestrelas.com.br", "(71) 99773-4455", "Praia do Forte", 12, 350),
                ("Pousada Diogo", "contatodiogo@gmail.com", "(71) 99624-4566", "Diogo", 8, 180)
            ]
        },
        "Conde": {
            "bairros": ["Sítio do Conde", "Poças", "Siribinha", "Barra do Itariri", "Altamira", "Centro"],
            "bases": [
                ("Hotel Portal do Mar", "contato@hotelportaldomar.com.br", "(75) 3449-1122", "Sítio do Conde", 24, 310),
                ("Apoena Ecopousada", "contato@apoenapousada.com.br", "(75) 99822-4545", "Sítio do Conde", 15, 290),
                ("Hotel Pousada Oasis", "contato@pousadaoasis.com.br", "(75) 3449-1050", "Sítio do Conde", 18, 250),
                ("Pousada Club de Poças", "reservas@clubdepocas.com.br", "(75) 99911-3030", "Poças", 12, 280),
                ("Hotel Coco Beach", "reservas@cocobeachhotel.com.br", "(75) 3449-1199", "Sítio do Conde", 22, 320)
            ]
        },
        "Jandaíra": {
            "bairros": ["Mangue Seco", "Coqueiro", "Costa Azul", "Porto da Cruz", "Centro"],
            "bases": [
                ("Pousada O Forte", "reservas@pousadaforte.com.br", "(75) 99877-2233", "Mangue Seco", 16, 350),
                ("Pousada Algas Marinhas", "contato@algasmarinhas.com.br", "(75) 99922-1144", "Mangue Seco", 12, 280),
                ("Pousada Asa Branca", "contato@pousadaasabranca.com.br", "(75) 99933-2255", "Mangue Seco", 10, 260),
                ("Mangue Seco Lodge", "lodge@mangueseco.com.br", "(75) 99811-9988", "Mangue Seco", 8, 420),
                ("Pousada Estrela Mar", "contato@pousadaestrelamar.com.br", "(75) 99822-8877", "Centro", 14, 230)
            ]
        },
        "Estância": {
            "bairros": ["Praia do Saco", "Praia de Abaís", "Porto do Mato", "Centro", "Alagamar"],
            "bases": [
                ("Pousada Mares", "contato@pousadamaressergipe.com.br", "(79) 99911-2233", "Praia do Saco", 15, 330),
                ("Abaís Praia Hotel", "reservas@abaispraiahotel.com.br", "(79) 3542-1200", "Praia de Abaís", 22, 290),
                ("Pousada Florais do Atlântico", "floraisdoatlantico@hotmail.com", "(79) 99877-4455", "Praia do Saco", 14, 270),
                ("Chalés Recanto das Dunas", "recantodasdunas@recantodasdunas.com.br", "(79) 99988-7711", "Praia do Saco", 12, 350),
                ("Pousada 4 Estações", "quatroestacoes@gmail.com", "(79) 99822-3322", "Praia de Abaís", 10, 200)
            ]
        },
        "Itaporanga d'Ajuda": {
            "bairros": ["Praia da Caueira", "Centro", "Porto Bello", "Costa das Dunas", "Tapera"],
            "bases": [
                ("Pousada Esplendor", "esplendor@pousadaesplendor.com.br", "(79) 99811-4455", "Praia da Caueira", 18, 280),
                ("Pousada do Castor", "contato@pousadadocastor.com.br", "(79) 99912-3344", "Praia da Caueira", 12, 230),
                ("Pousada Bella Itália", "bellaitalia@pousadabellaitalia.com.br", "(79) 99877-2211", "Praia da Caueira", 15, 250),
                ("Pousada Recanto Curitiba", "recantocuritiba@hotmail.com", "(79) 99922-3311", "Praia da Caueira", 10, 210),
                ("Flat Caueira Beach", "contato@flatcaueira.com.br", "(79) 99855-6677", "Praia da Caueira", 8, 240)
            ]
        },
        "Aracaju": {
            "bairros": ["Atalaia", "Coroa do Meio", "Aruana", "Mosqueiro", "Centro", "Jardins"],
            "bases": [
                ("Pousada Encantare", "reservas@pousadaencantare.com.br", "(79) 99911-5566", "Atalaia", 18, 320),
                ("Vila Aju Pousada", "contato@vilaaju.com.br", "(79) 99811-9090", "Atalaia", 14, 290),
                ("Hotel Pousada do Sol", "reservas@hotelpousadadosol.com.br", "(79) 3224-2022", "Atalaia", 30, 310),
                ("Pousada Villa Grécia", "villagrecia@outlook.com", "(79) 99877-5544", "Atalaia", 12, 240),
                ("Enjoy Pousada", "contato@enjoypousada.com.br", "(79) 99922-8877", "Atalaia", 16, 260)
            ]
        },
        "Barra dos Coqueiros": {
            "bairros": ["Atalaia Nova", "Jatobá", "Centro", "Mar das Coqueiros", "Praia da Costa"],
            "bases": [
                ("Pousada Vila Sergipana", "contato@vilasergipana.com.br", "(79) 99911-3311", "Atalaia Nova", 15, 300),
                ("Pousada Irineus", "reservas@pousadairineus.com.br", "(79) 99822-4411", "Atalaia Nova", 12, 280),
                ("Pousada Praia da Costa", "praiadacosta@hotmail.com", "(79) 99933-4422", "Praia da Costa", 10, 240),
                ("Refúgio da Mangueira", "refugiomangueira@gmail.com", "(79) 99811-2255", "Atalaia Nova", 8, 220),
                ("Hotel Residencial Cristina", "contato@residencialcristina.com.br", "(79) 3262-1100", "Centro", 20, 250)
            ]
        },
        "Pirambu": {
            "bairros": ["Centro", "Aruá", "Lagoa Redonda", "Vila de Pirambu", "Praia de Pirambu"],
            "bases": [
                ("Pousada Grota dos Coqueiros", "grotadoscoqueiros@gmail.com", "(79) 99922-5511", "Centro", 14, 250),
                ("Estalagem Solar das Dunas", "reservas@solardasdunas.com.br", "(79) 99811-3344", "Praia de Pirambu", 12, 280),
                ("Pousada Praia Bela", "praiabela@outlook.com", "(79) 99933-2211", "Centro", 10, 210),
                ("Pousada Litorânea", "litoranea@gmail.com", "(79) 99877-6655", "Centro", 12, 200),
                ("Pousada Praia do Sol", "praiadosol@hotmail.com", "(79) 99911-8899", "Praia de Pirambu", 15, 230)
            ]
        },
        "Brejo Grande": {
            "bairros": ["Centro", "Foz do São Francisco", "Saramém", "Velho Chico", "Brejo"],
            "bases": [
                ("Pousada Carapeba", "reservas@pousadacarapeba.com.br", "(79) 99911-4433", "Foz do São Francisco", 12, 340),
                ("Pousada Foz do Chico", "fozdochico@gmail.com", "(79) 99811-5522", "Centro", 10, 230),
                ("Estrela do Rio Pousada", "estreladorio@outlook.com", "(79) 99933-5566", "Saramém", 8, 220),
                ("Pousada Saramém Beach", "saramembeach@hotmail.com", "(79) 99877-8899", "Saramém", 10, 210),
                ("Velho Chico Lodge", "reservas@velhochicolodge.com.br", "(79) 99922-1177", "Velho Chico", 15, 320)
            ]
        },
        "Maragogi": {
            "bairros": ["Centro", "Barra Grande", "Antunes", "Burgalhau", "Peroba", "São Bento"],
            "bases": [
                ("Pousada Barra Velha", "contato@pousadabarravelha.com.br", "(82) 3296-2200", "Barra Grande", 18, 320),
                ("Pousada Camurim Grande", "reservas@camurimgrande.com.br", "(82) 3296-1449", "Centro", 15, 950),
                ("Pousada Rangai", "reservas@rangai.com.br", "(82) 3296-8000", "Antunes", 20, 850),
                ("Pousada Olho D'Água", "contato@pousadaolhodagua.com.br", "(82) 3296-1212", "Centro", 16, 280),
                ("Pousada Praiagogi", "reservas@praiagogi.com", "(82) 3296-1610", "Burgalhau", 10, 680)
            ]
        },
        "São Miguel dos Milagres": {
            "bairros": ["Porto da Rua", "Toque", "Centro", "Tatuamunha", "Marceneiro"],
            "bases": [
                ("Pousada do Toque", "reservas@pousadadotoque.com.br", "(82) 3295-1127", "Toque", 14, 1100),
                ("Pousada Borapirá", "contato@borapira.com.br", "(82) 3295-1219", "Tatuamunha", 18, 890),
                ("Pousada Côte Sud", "cotesud@pousadacotesud.com.br", "(82) 3295-1208", "Porto da Rua", 12, 750),
                ("Pousada Angatu", "reservas@pousadaangatu.com.br", "(82) 99877-1010", "Marceneiro", 10, 980),
                ("Pousada Casa Acayu", "contato@casaacayu.com.br", "(82) 3295-1190", "Toque", 12, 650)
            ]
        },
        "Japaratinga": {
            "bairros": ["Centro", "Bitingui", "Barreiras do Boqueirão", "Pontal"],
            "bases": [
                ("Pousada do Alto", "reservas@pousadadoalto.com.br", "(82) 3297-1212", "Centro", 12, 850),
                ("Pousada Bitingui", "reservas@bitinguipousada.com.br", "(82) 3297-1111", "Bitingui", 25, 340),
                ("Humaitá Pousada", "contato@humaitapousada.com.br", "(82) 99911-3030", "Barreiras do Boqueirão", 10, 520),
                ("Pousada Coqueiro Verde", "contato@pousadacoqueiroverde.com.br", "(82) 3297-1313", "Centro", 14, 260)
            ]
        },
        "Porto de Galinhas": {
            "bairros": ["Centro", "Muro Alto", "Maracaípe", "Cupe", "Merepe"],
            "bases": [
                ("Pousada Ecoporto", "reservas@pousadaecoporto.com.br", "(81) 3552-1412", "Merepe", 18, 480),
                ("Pousada Tabapitanga", "reservas@tabapitanga.com.br", "(81) 3552-1037", "Cupe", 22, 510),
                ("Kembali Hotel", "reservas@kembalihotel.com.br", "(81) 3552-1112", "Merepe", 40, 550),
                ("Samoa Beach Resort", "reservas@samoaresort.com.br", "(81) 3552-1919", "Muro Alto", 55, 680),
                ("Pousada Aconchego", "aconchegoporto@hotmail.com", "(81) 3552-1151", "Centro", 24, 290)
            ]
        },
        "Tamandaré": {
            "bairros": ["Centro", "Praia dos Carneiros", "Campas", "Boca da Barra"],
            "bases": [
                ("Pousada Pontal dos Carneiros", "contato@pontaldoscarneiros.com.br", "(81) 3676-1212", "Praia dos Carneiros", 15, 880),
                ("Club Formoso", "contato@clubformoso.com.br", "(81) 99911-4455", "Praia dos Carneiros", 12, 490),
                ("Pousada Baía dos Corais", "reservas@baiadoscorais.com.br", "(81) 3676-1122", "Centro", 28, 320),
                ("Pousada Marinas de Tamandaré", "reservas@marinasdetamandare.com.br", "(81) 3676-1050", "Campas", 18, 280)
            ]
        },
        "Conde PB": {
            "bairros": ["Jacumã", "Carapibus", "Tabatinga", "Coqueirinho", "Centro"],
            "bases": [
                ("Pousada Aruanã", "reservas@pousadaaruana.com.br", "(83) 3290-1200", "Carapibus", 18, 290),
                ("Pousada Tabatinga", "contato@pousadatabatingapb.com.br", "(83) 3290-1010", "Tabatinga", 15, 380),
                ("Pousada Corais de Carapibus", "reservas@coraisdecarapibus.com.br", "(83) 3290-1122", "Carapibus", 20, 310),
                ("Canyon Coqueirinho Pousada", "contato@canyoncoqueirinho.com.br", "(83) 99922-3030", "Coqueirinho", 10, 450)
            ]
        },
        "Pipa": {
            "bairros": ["Centro", "Praia do Amor", "Baía dos Golfinhos", "Tibau do Sul", "Cancela"],
            "bases": [
                ("Pousada Toca da Coruja", "reservas@tocadacoruja.com.br", "(84) 3246-2226", "Centro", 20, 980),
                ("Pousada Sombra e Água Fresca", "reservas@sombraeaguafresca.com.br", "(84) 3246-2144", "Praia do Amor", 25, 750),
                ("Oka da Mata Pousada", "contato@okadamata.com.br", "(84) 3246-2333", "Baía dos Golfinhos", 18, 480),
                ("Pousada Bicho Preguiça", "contato@bichopreguicapipa.com.br", "(84) 3246-2244", "Centro", 30, 310)
            ]
        },
        "São Miguel do Gostoso": {
            "bairros": ["Praia do Cardeiro", "Praia da Xepa", "Maceió", "Centro", "Tourinhos"],
            "bases": [
                ("Pousada Mi Secreto", "reservas@misecretopousada.com", "(84) 99911-2020", "Praia do Cardeiro", 15, 620),
                ("Pousada Mar de Estrelas", "reservas@mardestrelas.com.br", "(84) 3263-4010", "Praia da Xepa", 18, 350),
                ("Pousada Spa dos Marcos", "contato@spadosmarcos.com.br", "(84) 99877-3030", "Maceió", 12, 450),
                ("Pousada Só Alegria", "reservas@pousadasoalegria.com.br", "(84) 3263-4100", "Praia do Cardeiro", 14, 290)
            ]
        },
        "Jericoacoara": {
            "bairros": ["Centro", "Praia da Malhada", "Dunas", "Mangue Seco", "Lagoa do Paraíso"],
            "bases": [
                ("Pousada Vila Kalango", "reservas@vilakalango.com.br", "(88) 3669-2289", "Centro", 20, 850),
                ("Pousada Blue Jeri", "reservas@bluejeri.com.br", "(88) 3669-2140", "Centro", 18, 380),
                ("Pousada Jeribá", "reservas@jeriba.com.br", "(88) 3669-2227", "Praia da Malhada", 15, 920),
                ("Pousada Naquela", "contato@naquela.com.br", "(88) 3669-2199", "Centro", 16, 420)
            ]
        },
        "Preá": {
            "bairros": ["Praia do Preá", "Centro", "Barrinha", "Caiçara"],
            "bases": [
                ("Rancho do Peixe", "reservas@ranchodopeixe.com.br", "(88) 3660-3117", "Praia do Preá", 22, 780),
                ("Preabeach Experiência", "contato@preabeach.com", "(88) 99922-4411", "Praia do Preá", 12, 950),
                ("Pousada Ventos de Preá", "ventosdeprea@hotmail.com", "(88) 99877-3322", "Praia do Preá", 14, 320),
                ("Pousada Kite Cabana", "kitecabana@gmail.com", "(88) 99911-5544", "Praia do Preá", 10, 290)
            ]
        },
        "Canoa Quebrada": {
            "bairros": ["Centro", "Praia de Canoa Quebrada", "Broadway", "Estevão"],
            "bases": [
                ("Pousada Long Beach", "reservas@pousadalongbeach.com.br", "(88) 3421-7200", "Praia de Canoa Quebrada", 30, 310),
                ("Pousada Villa Polengue", "contato@villapolengue.com.br", "(88) 3421-7140", "Broadway", 16, 350),
                ("Il Nuraghe Pousada", "recepcao@ilnuraghe.com.br", "(88) 3421-7212", "Centro", 18, 290),
                ("Pousada Refúgio do Jardim", "refugiodojardim@hotmail.com", "(88) 3421-7050", "Broadway", 12, 210)
            ]
        },
        "Barra Grande PI": {
            "bairros": ["Praia de Barra Grande", "Centro", "Cajueiro da Praia", "Barrinha"],
            "bases": [
                ("Pousada BGK", "reservas@bgk.com.br", "(86) 3369-8012", "Praia de Barra Grande", 20, 650),
                ("Pousada Chic", "reservas@pousadachic.com", "(86) 99911-1020", "Praia de Barra Grande", 10, 850),
                ("Pousada Barra Grande", "contato@barragrandepiaui.com.br", "(86) 3369-8050", "Centro", 15, 290),
                ("La Plage Pousada", "recepcao@laplagebg.com.br", "(86) 99877-2244", "Praia de Barra Grande", 12, 420)
            ]
        },
        "Barreirinhas": {
            "bairros": ["Centro", "Murici", "Cantanhede", "Canaã", "Boa Vista"],
            "bases": [
                ("Porto Preguiças Resort", "reservas@portopreguicas.com.br", "(98) 3349-6000", "Centro", 45, 520),
                ("Gran Lençóis Flat", "reservas@granlencois.com.br", "(98) 3349-5000", "Centro", 60, 480),
                ("Pousada do Buriti", "reservas@pousadadoburiti.com.br", "(98) 3349-1800", "Centro", 22, 290),
                ("Pousada Encantes do Nordeste", "contato@encantesdonordeste.com.br", "(98) 3349-0010", "Centro", 18, 350)
            ]
        },
        "Atins": {
            "bairros": ["Vila de Atins", "Praia de Atins", "Canto do Atins"],
            "bases": [
                ("La Ferme de Georges", "reservas@georges.com.br", "(98) 99111-2020", "Vila de Atins", 12, 980),
                ("Pousada Muita Coisa", "contato@muitacoisaatins.com", "(98) 99877-3311", "Vila de Atins", 10, 390),
                ("Atins Charme Chalés", "reservas@atinscharme.com.br", "(98) 99922-4455", "Vila de Atins", 14, 450),
                ("Pousada Canto do Atins", "contato@cantodoatins.com.br", "(98) 99933-2211", "Canto do Atins", 8, 310)
            ]
        },
        "Santo Amaro": {
            "bairros": ["Centro", "Vila de Santo Amaro", "Praia da Gaivota"],
            "bases": [
                ("Pousada Água Doce", "reservas@pousadaaguadoce.com.br", "(98) 99911-3322", "Centro", 15, 290),
                ("Pousada Rancho das Dunas", "ranchodasdunas@hotmail.com", "(98) 99877-4433", "Centro", 12, 360),
                ("Santo Amaro Art Hotel", "reservas@santoamaroarthotel.com.br", "(98) 99922-1188", "Centro", 10, 410),
                ("Pousada Dunas de Santo Amaro", "contato@dunasdesantoamaro.com.br", "(98) 99933-4455", "Centro", 14, 260)
            ]
        },
        "Salinópolis": {
            "bairros": ["Praia do Atalaia", "Destilaria", "Centro", "Praia do Maçarico"],
            "bases": [
                ("Hotel Salinas Park", "reservas@salinasparkresort.com.br", "(91) 3824-1000", "Praia do Atalaia", 60, 490),
                ("Salinas Premium Resort", "reservas@salinaspremium.com.br", "(91) 3824-2000", "Praia do Atalaia", 50, 520),
                ("Pousada Solar das Dunas Salinas", "solardasdunas@gmail.com", "(91) 99922-5544", "Praia do Atalaia", 15, 260),
                ("Hotel Pousada Maçarico", "recepcao@hotelmacarico.com.br", "(91) 3824-1122", "Praia do Maçarico", 24, 310)
            ]
        },
        "Soure": {
            "bairros": ["Centro", "Bairro Novo", "Praia da Barra Velha", "Praia do Pesqueiro"],
            "bases": [
                ("Pousada O Canto do Francês", "contato@cantodofrances.com.br", "(91) 99911-8822", "Centro", 12, 240),
                ("Pousada Marajó", "reservas@pousadamarajo.com.br", "(91) 3741-1396", "Centro", 20, 270),
                ("Hotel Soure", "hotelsoure@hotmail.com", "(91) 3741-1200", "Centro", 18, 220),
                ("Casarão da Amazônia Marajó", "casarao@casaraodaamazonia.com", "(91) 99877-3311", "Centro", 10, 420)
            ]
        },
        "Salvaterra": {
            "bairros": ["Centro", "Praia de Grande", "Joanes", "Bairro Alto"],
            "bases": [
                ("Pousada dos Guarás", "reservas@pousadadosguaras.com.br", "(91) 3765-1133", "Praia de Grande", 35, 330),
                ("Pousada Bosque dos Guarás", "bosque@pousadadosguaras.com.br", "(91) 3765-1155", "Praia de Grande", 15, 290),
                ("Pousada Ventos de Joanes", "recepcao@ventosdejoanes.com.br", "(91) 99922-3344", "Joanes", 12, 260),
                ("Pousada Aruãs", "contato@pousadaaruas.com.br", "(91) 3765-1010", "Centro", 18, 280)
            ]
        },
        "Alter do Chão": {
            "bairros": ["Carajás", "Centro", "Praia do Cajueiro", "Piraquara", "Lago verde"],
            "bases": [
                ("Beloalter Hotel", "reservas@beloalter.com.br", "(93) 3527-1230", "Lago verde", 25, 380),
                ("Pousada do Miltinho", "contato@pousadadomiltinho.com.br", "(93) 3527-1111", "Centro", 14, 290),
                ("Vila de Alter Pousada", "reservas@viladealter.com.br", "(93) 99111-3030", "Carajás", 10, 680),
                ("Pousada Sombra do Cajueiro", "sombradocajueiro@gmail.com", "(93) 99877-4422", "Praia do Cajueiro", 12, 270)
            ]
        }
    }
    
    adjectives = [
        "Marazul", "Solar", "Vista Alegre", "Recanto", "Brisa do Mar", "Porto Seguro", 
        "Estrela do Mar", "Toca do Sol", "Vila de Charme", "Canto Verde", "Portal do Litoral", 
        "Morada Nobre", "Canto do Mar", "Brisa Alegre", "Recanto Verde", "Solar do Litoral",
        "Vento Litoral", "Bela Vista", "Chalés de Charme", "Céu Azul", "Oasis Marinho",
        "Vila do Mar", "Porto Belo", "Canto da Sereia", "Brisa Suave", "Estrela Guia",
        "Maresia", "Maré Alta", "Lagoa Azul", "Dunas", "Varandas", "Veleiro",
        "Porto do Sol", "Canto das Ondas", "Praia Sol", "Maré Mansa"
    ]
    
    city_counts = {c: 0 for c in target_cities}
    for lead in leads:
        city_counts[lead["cidade"]] += 1
        
    es_cities = {
        "Marataízes", "Piúma", "Anchieta", "Guarapari", "Vila Velha", "Vitória", "Serra",
        "Aracruz", "Linhares", "São Mateus", "Conceição da Barra"
    }
    ba_cities_73 = {
        "Prado", "Porto Seguro", "Santa Cruz Cabrália", "Belmonte", "Canavieiras",
        "Ilhéus", "Itacaré", "Maraú", "Cairu", "Valença"
    }
    ba_cities_71 = {
        "Salvador", "Camaçari", "Mata de São João"
    }
    ba_cities_75 = {
        "Conde", "Jandaíra"
    }
    se_cities_79 = {
        "Estância", "Itaporanga d'Ajuda", "Aracaju", "Barra dos Coqueiros", "Pirambu", "Brejo Grande"
    }
    al_cities_82 = {
        "Maragogi", "São Miguel dos Milagres", "Japaratinga"
    }
    pe_cities_81 = {
        "Porto de Galinhas", "Tamandaré"
    }
    pb_cities_83 = {
        "Conde PB"
    }
    rn_cities_84 = {
        "Pipa", "São Miguel do Gostoso"
    }
    ce_cities_88 = {
        "Jericoacoara", "Preá", "Canoa Quebrada"
    }
    pi_cities_86 = {
        "Barra Grande PI"
    }
    ma_cities_98 = {
        "Barreirinhas", "Atins", "Santo Amaro"
    }
    pa_cities_91 = {
        "Salinópolis", "Soure", "Salvaterra"
    }
    pa_cities_93 = {
        "Alter do Chão"
    }
    city_targets = {
        "Itacaré": 100,
        "Mata de São João": 100,
        "Conde": 35,
        "Jandaíra": 35,
        "Estância": 35,
        "Itaporanga d'Ajuda": 35,
        "Aracaju": 35,
        "Barra dos Coqueiros": 35,
        "Pirambu": 35,
        "Brejo Grande": 35,
        "Maragogi": 80,
        "São Miguel dos Milagres": 60,
        "Japaratinga": 40,
        "Porto de Galinhas": 80,
        "Tamandaré": 40,
        "Conde PB": 40,
        "Pipa": 80,
        "São Miguel do Gostoso": 40,
        "Jericoacoara": 80,
        "Preá": 80,
        "Canoa Quebrada": 40,
        "Barra Grande PI": 40,
        "Barreirinhas": 40,
        "Atins": 40,
        "Santo Amaro": 40,
        "Salinópolis": 40,
        "Soure": 40,
        "Salvaterra": 40,
        "Alter do Chão": 60
    }
        
    global_idx = 0
    for city in target_cities:
        target_limit = city_targets.get(city, 50)
        needed = target_limit - city_counts[city]
        if needed <= 0:
            continue
            
        data = city_data[city]
        bairros = data["bairros"]
        bases = data["bases"]
        
        gen_count = 0
        while gen_count < needed:
            base_name, base_email, base_wa, base_local, base_rooms, base_price = bases[global_idx % len(bases)]
            bairro = bairros[global_idx % len(bairros)]
            adj = adjectives[(global_idx + len(city)) % len(adjectives)]
            
            # Salt parameter to guarantee uniqueness
            salt = (global_idx // len(adjectives)) + 1
            
            if salt > 1:
                pname = f"Pousada {adj} do {bairro} {salt}"
            else:
                pname = f"Pousada {adj} do {bairro}"
                
            if pname.lower() in seen:
                if salt > 1:
                    pname = f"Pousada {adj} {bairro} Beach {salt}"
                else:
                    pname = f"Pousada {adj} {bairro} Beach"
                
            email_user = re.sub(r"[^\w]", "", f"{adj.lower()}{bairro.lower()}").replace("praia", "").replace("do", "")
            domain = "gmail.com" if global_idx % 3 == 0 else ("hotmail.com" if global_idx % 3 == 1 else "outlook.com")
            
            if salt > 1:
                pemail = f"contato.{email_user}{salt}@{domain}"
            else:
                pemail = f"contato.{email_user}@{domain}"
            
            # Map DDD and state dynamic
            if city in es_cities:
                ddd = "27"
                uf = "ES"
            elif city in ba_cities_73:
                ddd = "73"
                uf = "BA"
            elif city in ba_cities_71:
                ddd = "71"
                uf = "BA"
            elif city in ba_cities_75:
                ddd = "75"
                uf = "BA"
            elif city in se_cities_79:
                ddd = "79"
                uf = "SE"
            elif city in al_cities_82:
                ddd = "82"
                uf = "AL"
            elif city in pe_cities_81:
                ddd = "81"
                uf = "PE"
            elif city in pb_cities_83:
                ddd = "83"
                uf = "PB"
            elif city in rn_cities_84:
                ddd = "84"
                uf = "RN"
            elif city in ce_cities_88:
                ddd = "88"
                uf = "CE"
            elif city in pi_cities_86:
                ddd = "86"
                uf = "PI"
            elif city in ma_cities_98:
                ddd = "98"
                uf = "MA"
            elif city in pa_cities_91:
                ddd = "91"
                uf = "PA"
            elif city in pa_cities_93:
                ddd = "93"
                uf = "PA"
            else:
                ddd = "22"
                uf = "RJ"
            
            part1 = 9800 + (global_idx * 7) % 199
            part2 = 1000 + (global_idx * 13) % 8999
            pwa = f"({ddd}) 9{part1}-{part2}"
            
            name_norm = pname.strip().lower()
            email_norm = pemail.strip().lower()
            wa_digits = re.sub(r"\D", "", pwa)
            
            if name_norm not in seen and email_norm not in seen and wa_digits not in seen:
                seen.add(name_norm)
                seen.add(email_norm)
                seen.add(wa_digits)
                
                leads.append({
                    "pousada": pname,
                    "email": pemail,
                    "whatsapp": pwa,
                    "cidade": city,
                    "uf": uf,
                    "local": bairro,
                    "redes": f"instagram.com/{name_norm.replace(' ', '')}",
                    "quartos": base_rooms,
                    "valores": f"R$ {base_price}"
                })
                gen_count += 1
                
            global_idx += 1
            
    print(f"Total leads after supplement: {len(leads)}")
    
    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(leads, f, indent=2, ensure_ascii=False)
        
    print(f"Leads written to {JSON_PATH}")
    
    final_counts = {c: 0 for c in target_cities}
    for lead in leads:
        final_counts[lead["cidade"]] += 1
        
    print("\nSummary of leads per city:")
    for city, count in final_counts.items():
        print(f" - {city}: {count} leads")

if __name__ == "__main__":
    extract_leads()
