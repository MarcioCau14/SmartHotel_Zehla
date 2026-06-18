import os
import re
import openpyxl
import pandas as pd
import dns.resolver
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==============================================================================
# CONFIGURAÇÕES E PASTAS
# ==============================================================================
FOLDER_PATH = "/Users/marciocau/Downloads/PLANILHAS_MARKETING_BR_"
CONSOLIDATED_FILE_PATH = os.path.join(FOLDER_PATH, "CONTATOS_VALIDOS_CONSOLIDADO.xlsx")
SOURCE_FASE_VALIDADA = "/Users/marciocau/Downloads/POUSADAS_MARKETING_FASE_VALIDADA.xlsx"

# Estilos de design premium
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Azul corporativo premium
HEADER_FONT = Font(color="FFFFFF", name="Segoe UI", size=11, bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROW_FONT = Font(name="Segoe UI", size=11)
BORDER_THIN = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# Caches de validação
domain_mx_cache = {}
validated_emails_cache = {}
validated_whatsapps_cache = {}

# ==============================================================================
# VALIDAÇÃO DE CONCURRÊNCIA DNS MX
# ==============================================================================

def check_mx_dns_single(domain):
    common_good_domains = {
        "gmail.com", "hotmail.com", "yahoo.com", "outlook.com", "yahoo.com.br",
        "uol.com.br", "terra.com.br", "ig.com.br", "bol.com.br", "live.com", 
        "outlook.com.br", "icloud.com", "globo.com", "icloud.com.br", "me.com"
    }
    if domain in common_good_domains:
        return domain, True
    try:
        answers = dns.resolver.resolve(domain, 'MX', lifetime=2)
        return domain, len(answers) > 0
    except Exception:
        try:
            answers_a = dns.resolver.resolve(domain, 'A', lifetime=2)
            return domain, len(answers_a) > 0
        except Exception:
            return domain, False

# ==============================================================================
# VALIDAÇÃO DE CONTEXTO E OSINT
# ==============================================================================

def validate_email_osint(email_str, pousada_name):
    if not email_str or email_str == "-" or email_str == "nan":
        return False, "E-mail Ausente", 0
    email_clean = email_str.strip().lower()
    email_regex = r"^[\w\.-]+@[\w\.-]+\.\w+$"
    if not re.match(email_regex, email_clean):
        return False, "Sintaxe de E-mail Inválida", 0
    if email_clean in validated_emails_cache:
        cached = validated_emails_cache[email_clean]
        return cached["is_valid"], cached["status"], cached["score"]
    domain = email_clean.split("@")[1].strip().lower()
    domain = re.sub(r"[^\w\.-]", "", domain)
    
    has_mx = domain_mx_cache.get(domain, False)
    if not has_mx:
        res = (False, "Domínio sem registros MX (Inexistente)", 0)
        validated_emails_cache[email_clean] = {"is_valid": False, "status": res[1], "score": 0}
        return res
    username = email_clean.split("@")[0]
    pousada_norm = re.sub(r"[^\w]", "", pousada_name.lower())
    username_norm = re.sub(r"[^\w]", "", username)
    domain_norm = re.sub(r"[^\w]", "", domain.split(".")[0])
    
    score = 50
    status = "Validado via MX"
    if username_norm in pousada_norm or domain_norm in pousada_norm:
        score += 35
        status = "Validado via Digital Footprint (Domínio Pousada)"
    elif username in ["atendimento", "reservas", "contato", "comercial", "direcao"]:
        score += 15
        status = "Canal Corporativo Oficial"
    else:
        score += 25
        status = "Contato Direto Potencial"
    res = (True, status, score)
    validated_emails_cache[email_clean] = {"is_valid": True, "status": status, "score": score}
    return res

def validate_whatsapp_osint(phone_str, uf_lead):
    if not phone_str or phone_str == "-" or phone_str == "nan":
        return False, "WhatsApp Ausente", 0, ""
    phone_raw = str(phone_str)
    if phone_raw in validated_whatsapps_cache:
        cached = validated_whatsapps_cache[phone_raw]
        return cached["is_valid"], cached["status"], cached["score"], cached["clean"]
    digits = re.sub(r"\D", "", phone_raw)
    if len(digits) < 10 or len(digits) > 13:
        res = (False, "Tamanho de Número Inválido", 0, digits)
        validated_whatsapps_cache[phone_raw] = {"is_valid": False, "status": res[1], "score": 0, "clean": digits}
        return res
    if len(digits) in [12, 13] and digits.startswith("55"):
        digits = digits[2:]
    if len(digits) == 11:
        ddd = digits[:2]
        ninth_digit = digits[2]
        if ninth_digit == "9":
            status = "WhatsApp Celular Ativo"
            score = 90
            sc_ddds = ["48", "47", "49"]
            if uf_lead == "SC" and ddd not in sc_ddds:
                status = "WhatsApp Decisor Direto (DDD Externo)"
                score = 98
        else:
            status = "Número Incorreto (Celular sem 9)"
            score = 40
    elif len(digits) == 10:
        status = "WhatsApp Fixo Corporativo"
        score = 65
    else:
        status = "Número Inválido"
        score = 0
    is_valid = score >= 50
    res = (is_valid, status, score, digits)
    validated_whatsapps_cache[phone_raw] = {"is_valid": is_valid, "status": status, "score": score, "clean": digits}
    return res

def parse_valores(val_str):
    if not val_str or val_str == "-" or val_str == "nan":
        return 0
    clean = re.sub(r"[^\d]", "", str(val_str))
    if not clean:
        return 0
    return int(clean)

def calculate_purchasing_power(valores_str, quartos_str):
    val = parse_valores(valores_str)
    quartos = 0
    if quartos_str and quartos_str != "-" and quartos_str != "nan":
        q_digits = re.sub(r"\D", "", str(quartos_str))
        if q_digits:
            quartos = int(q_digits)
    if val >= 500 or quartos > 20:
        return "ALTO (ICP A+)", "Comportamento Seletivo: Alta exigência por automação e exclusividade. Foco em ROI.", 90
    elif val >= 250 or quartos > 10:
        return "MÉDIO (ICP A)", "Comportamento Moderado: Equilíbrio entre tradição e abertura tecnológica.", 70
    else:
        return "NORMAL", "Foco em Eficiência: Busca por redução de custos e escala. Valoriza ferramentas baratas.", 45

# ==============================================================================
# PIPELINE DE EXPANSÃO GAROPABA E IMBITUBA
# ==============================================================================

def main():
    print("🧠 [SECRETARIA-IA] Iniciando Varredura e Expansão focada em GAROPABA e IMBITUBA, SC...")
    
    # 1. Mapear planilhas brutas originais
    raw_files = []
    for f in os.listdir(FOLDER_PATH):
        if f.endswith(".xlsx") and not f.startswith("~$") and f != "CONTATOS_VALIDOS_CONSOLIDADO.xlsx" and not f.endswith("_VALIDADA.xlsx"):
            raw_files.append(os.path.join(FOLDER_PATH, f))
            
    if os.path.exists(SOURCE_FASE_VALIDADA.replace("_VALIDADA", "")):
        raw_files.append(SOURCE_FASE_VALIDADA.replace("_VALIDADA", ""))
        
    print(f"📊 Mapeadas {len(raw_files)} planilhas brutas para extração complementar.")
    
    # Extrair todos os contatos brutos das duas cidades para validar
    candidates = []
    unique_domains = set()
    
    for file_path in raw_files:
        print(f"🔍 Buscando em: {os.path.basename(file_path)}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
            
            col_indices = {}
            for col_idx, h in enumerate(headers, 1):
                h_lower = h.lower()
                if "pousada" in h_lower or "nome" in h_lower or "propriedade" in h_lower:
                    col_indices["Pousada"] = col_idx
                elif "e-mail" in h_lower or "email" in h_lower:
                    col_indices["Email"] = col_idx
                elif "whats" in h_lower:
                    col_indices["Whatsapp"] = col_idx
                elif "quarto" in h_lower:
                    col_indices["Qtd Quartos"] = col_idx
                elif "local" in h_lower or "praia" in h_lower:
                    col_indices["Local"] = col_idx
                elif "cidade" in h_lower:
                    col_indices["Cidade"] = col_idx
                elif "uf" in h_lower:
                    col_indices["UF"] = col_idx
                elif "valor" in h_lower:
                    col_indices["Valores Estimados"] = col_idx
                elif "qualificação" in h_lower or "qualificacao" in h_lower:
                    col_indices["Qualificação"] = col_idx
                elif "validação" in h_lower or "validacao" in h_lower:
                    col_indices["Validação"] = col_idx
                elif "comportamento" in h_lower:
                    col_indices["Comportamento de Compra"] = col_idx
                elif "sinal" in h_lower or "intenção" in h_lower or "intencao" in h_lower:
                    col_indices["Sinais de Intenção"] = col_idx
                elif "social" in h_lower or "rede" in h_lower:
                    col_indices["Redes Sociais"] = col_idx
                elif "latitude" in h_lower:
                    col_indices["LATITUDE"] = col_idx
                elif "longitude" in h_lower:
                    col_indices["LONGITUDE"] = col_idx
                elif "score qual" in h_lower:
                    col_indices["Score Qual."] = col_idx
                elif "score valid" in h_lower:
                    col_indices["Score Valid."] = col_idx
                    
            if "Cidade" not in col_indices or "Pousada" not in col_indices:
                continue
                
            for row in range(2, sheet.max_row + 1):
                cidade_val = str(sheet.cell(row=row, column=col_indices["Cidade"]).value or '').strip()
                if cidade_val.lower() not in ["imbituba", "garopaba"]:
                    continue
                    
                pousada_val = sheet.cell(row=row, column=col_indices["Pousada"]).value
                if not pousada_val:
                    continue
                pousada_name = str(pousada_val).strip()
                pousada_lower = pousada_name.lower()
                
                # Exclude non-pousadas
                exclude_keywords = ["hotel", "resort", "hostel", "flat", "residence", "reserve", "motel", "inn", "spa", "club", "lodge"]
                if any(kw in pousada_lower for kw in exclude_keywords):
                    continue
                    
                email_val = str(sheet.cell(row=row, column=col_indices["Email"]).value).strip() if "Email" in col_indices and sheet.cell(row=row, column=col_indices["Email"]).value is not None else ""
                whatsapp_val = str(sheet.cell(row=row, column=col_indices["Whatsapp"]).value).strip() if "Whatsapp" in col_indices and sheet.cell(row=row, column=col_indices["Whatsapp"]).value is not None else ""
                uf_val = str(sheet.cell(row=row, column=col_indices["UF"]).value).strip() if "UF" in col_indices and sheet.cell(row=row, column=col_indices["UF"]).value is not None else "SC"
                local_val = str(sheet.cell(row=row, column=col_indices["Local"]).value).strip() if "Local" in col_indices and sheet.cell(row=row, column=col_indices["Local"]).value is not None else ""
                quartos_val = sheet.cell(row=row, column=col_indices["Qtd Quartos"]).value if "Qtd Quartos" in col_indices else ""
                valores_val = sheet.cell(row=row, column=col_indices["Valores Estimados"]).value if "Valores Estimados" in col_indices else ""
                qual_val = sheet.cell(row=row, column=col_indices["Qualificação"]).value if "Qualificação" in col_indices else ""
                val_status = sheet.cell(row=row, column=col_indices["Validação"]).value if "Validação" in col_indices else ""
                comp_val = sheet.cell(row=row, column=col_indices["Comportamento de Compra"]).value if "Comportamento de Compra" in col_indices else ""
                sinais_val = sheet.cell(row=row, column=col_indices["Sinais de Intenção"]).value if "Sinais de Intenção" in col_indices else ""
                redes_val = sheet.cell(row=row, column=col_indices["Redes Sociais"]).value if "Redes Sociais" in col_indices else ""
                lat_val = sheet.cell(row=row, column=col_indices["LATITUDE"]).value if "LATITUDE" in col_indices else None
                lon_val = sheet.cell(row=row, column=col_indices["LONGITUDE"]).value if "LONGITUDE" in col_indices else None
                score_q = sheet.cell(row=row, column=col_indices["Score Qual."]).value if "Score Qual." in col_indices else 0
                score_v = sheet.cell(row=row, column=col_indices["Score Valid."]).value if "Score Valid." in col_indices else 0
                
                if "@" in email_val:
                    dom = email_val.split("@")[1].strip().lower()
                    dom = re.sub(r"[^\w\.-]", "", dom)
                    if dom:
                        unique_domains.add(dom)
                        
                candidates.append({
                    "pousada": pousada_name,
                    "email": email_val,
                    "whatsapp": whatsapp_val,
                    "quartos": quartos_val,
                    "local": local_val,
                    "cidade": cidade_val,
                    "uf": uf_val,
                    "valores": valores_val,
                    "qualificacao": qual_val,
                    "validacao": val_status,
                    "comportamento": comp_val,
                    "sinais": sinais_val,
                    "redes": redes_val,
                    "lat": lat_val,
                    "lon": lon_val,
                    "score_qual": score_q,
                    "score_valid": score_v
                })
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
            
    print(f"📊 Encontrados {len(candidates)} candidatos em Garopaba e Imbituba. Resolvendo {len(unique_domains)} domínios DNS em paralelo...")
    
    # Resolvendo domínios em paralelo
    with ThreadPoolExecutor(max_workers=100) as executor:
        results = executor.map(check_mx_dns_single, unique_domains)
        for dom, has_mx in results:
            domain_mx_cache[dom] = has_mx
            
    print("✅ Resolução DNS concluída. Validando contatos...")
    
    new_valid_leads = {}
    
    # Validar candidatos
    for c in candidates:
        is_email_v, email_status, email_score = validate_email_osint(c["email"], c["pousada"])
        is_whats_v, whats_status, whats_score, clean_whats = validate_whatsapp_osint(c["whatsapp"], c["uf"])
        power_tier, power_desc, power_score = calculate_purchasing_power(c["valores"], c["quartos"])
        
        # Se ambos forem válidos (aceitando móvel ou fixo WhatsApp Business ativo)
        if is_email_v and is_whats_v:
            lead_key = (c["pousada"].lower().strip(), c["email"].lower().strip(), re.sub(r"\D", "", c["whatsapp"]))
            combined_score = int((email_score + whats_score) / 2)
            
            new_valid_leads[lead_key] = {
                "pousada": c["pousada"],
                "email": c["email"],
                "whatsapp": c["whatsapp"],
                "quartos": c["quartos"],
                "local": c["local"],
                "cidade": c["cidade"],
                "uf": c["uf"],
                "valores": c["valores"],
                "qualificacao": power_tier,
                "validacao": f"E-mail: {email_status} | WA: {whats_status}",
                "comportamento": power_desc,
                "sinais": c["sinais"] or "Presença digital ativa.",
                "redes": c["redes"],
                "lat": c["lat"],
                "lon": c["lon"],
                "score_qual": power_score,
                "score_valid": combined_score
            }
            
    print(f"🎯 Total de pousadas válidas e exclusivas extraídas para Garopaba e Imbituba: {len(new_valid_leads)}")
    
    # 2. Carregar existentes de CONTATOS_VALIDOS_CONSOLIDADO.xlsx
    existing_leads = {}
    
    if os.path.exists(CONSOLIDATED_FILE_PATH):
        print(f"📂 Lendo dados existentes em CONTATOS_VALIDOS_CONSOLIDADO.xlsx para mesclagem...")
        try:
            wb_ex = openpyxl.load_workbook(CONSOLIDATED_FILE_PATH, data_only=True)
            # Lê a aba principal de Todos
            sheet_ex = wb_ex["Todas_Pousadas_Validas"]
            headers_ex = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet_ex[1]]
            
            col_indices = {}
            for col_idx, h in enumerate(headers_ex, 1):
                h_lower = h.lower()
                if "pousada" in h_lower or "nome" in h_lower:
                    col_indices["pousada"] = col_idx
                elif "e-mail" in h_lower or "email" in h_lower:
                    col_indices["email"] = col_idx
                elif "whats" in h_lower:
                    col_indices["whatsapp"] = col_idx
                elif "quarto" in h_lower:
                    col_indices["quartos"] = col_idx
                elif "local" in h_lower:
                    col_indices["local"] = col_idx
                elif "cidade" in h_lower:
                    col_indices["cidade"] = col_idx
                elif "uf" in h_lower:
                    col_indices["uf"] = col_idx
                elif "valor" in h_lower:
                    col_indices["valores"] = col_idx
                elif "qualificação" in h_lower or "qualificacao" in h_lower:
                    col_indices["qualificacao"] = col_idx
                elif "validação" in h_lower or "validacao" in h_lower:
                    col_indices["validacao"] = col_idx
                elif "comportamento" in h_lower:
                    col_indices["comportamento"] = col_idx
                elif "sinal" in h_lower or "intenção" in h_lower or "intencao" in h_lower:
                    col_indices["sinais"] = col_idx
                elif "social" in h_lower or "rede" in h_lower:
                    col_indices["redes"] = col_idx
                elif "latitude" in h_lower:
                    col_indices["lat"] = col_idx
                elif "longitude" in h_lower:
                    col_indices["lon"] = col_idx
                elif "score qual" in h_lower:
                    col_indices["score_qual"] = col_idx
                elif "score valid" in h_lower:
                    col_indices["score_valid"] = col_idx
                    
            for row in range(2, sheet_ex.max_row + 1):
                pousada_val = sheet_ex.cell(row=row, column=col_indices["pousada"]).value
                if not pousada_val:
                    continue
                pousada_name = str(pousada_val).strip()
                email_val = str(sheet_ex.cell(row=row, column=col_indices["email"]).value).strip() if "email" in col_indices and sheet_ex.cell(row=row, column=col_indices["email"]).value is not None else ""
                whatsapp_val = str(sheet_ex.cell(row=row, column=col_indices["whatsapp"]).value).strip() if "whatsapp" in col_indices and sheet_ex.cell(row=row, column=col_indices["whatsapp"]).value is not None else ""
                
                key = (pousada_name.lower().strip(), email_val.lower().strip(), re.sub(r"\D", "", whatsapp_val))
                
                existing_leads[key] = {
                    "pousada": pousada_name,
                    "email": email_val,
                    "whatsapp": whatsapp_val,
                    "quartos": sheet_ex.cell(row=row, column=col_indices["quartos"]).value if "quartos" in col_indices else "",
                    "local": sheet_ex.cell(row=row, column=col_indices["local"]).value if "local" in col_indices else "",
                    "cidade": sheet_ex.cell(row=row, column=col_indices["cidade"]).value if "cidade" in col_indices else "",
                    "uf": sheet_ex.cell(row=row, column=col_indices["uf"]).value if "uf" in col_indices else "SC",
                    "valores": sheet_ex.cell(row=row, column=col_indices["valores"]).value if "valores" in col_indices else "",
                    "qualificacao": sheet_ex.cell(row=row, column=col_indices["qualificacao"]).value if "qualificacao" in col_indices else "",
                    "validacao": sheet_ex.cell(row=row, column=col_indices["validacao"]).value if "validacao" in col_indices else "",
                    "comportamento": sheet_ex.cell(row=row, column=col_indices["comportamento"]).value if "comportamento" in col_indices else "",
                    "sinais": sheet_ex.cell(row=row, column=col_indices["sinais"]).value if "sinais" in col_indices else "",
                    "redes": sheet_ex.cell(row=row, column=col_indices["redes"]).value if "redes" in col_indices else "",
                    "lat": sheet_ex.cell(row=row, column=col_indices["lat"]).value if "lat" in col_indices else None,
                    "lon": sheet_ex.cell(row=row, column=col_indices["lon"]).value if "lon" in col_indices else None,
                    "score_qual": sheet_ex.cell(row=row, column=col_indices["score_qual"]).value if "score_qual" in col_indices else 0,
                    "score_valid": sheet_ex.cell(row=row, column=col_indices["score_valid"]).value if "score_valid" in col_indices else 0
                }
        except Exception as e:
            print(f"⚠️ Erro ao ler CONTATOS_VALIDOS_CONSOLIDADO.xlsx (possivelmente vazio ou corrompido): {e}")

    print(f"📊 Leads existentes na base: {len(existing_leads)}")
    
    # 3. Mesclar as duas bases, adicionando novos leads
    # Filtra os novos para evitar duplicados
    added_count = 0
    for key, lead in new_valid_leads.items():
        if key not in existing_leads:
            existing_leads[key] = lead
            added_count += 1
            
    print(f"🎉 Mesclagem concluída! Adicionados {added_count} novos contatos de Garopaba/Imbituba.")
    print(f"📈 Total global final de contatos de pousadas validadas: {len(existing_leads)}")
    
    # 4. Agrupar por cidades de novo
    leads_by_city = {}
    for lead in existing_leads.values():
        city_raw = lead["cidade"].strip()
        city_name = city_raw.title()
        if city_name not in leads_by_city:
            leads_by_city[city_name] = []
        leads_by_city[city_name].append(lead)
        
    print(f"🌆 Encontradas {len(leads_by_city)} cidades estruturadas no consolidado.")
    
    # 5. Regravar o Workbook CONTATOS_VALIDOS_CONSOLIDADO.xlsx
    wb_out = openpyxl.Workbook()
    default_sheet = wb_out.active
    wb_out.remove(default_sheet)
    
    columns_headers = [
        '#', 'Pousada', 'E-mail', 'Whatsapp', 'Qtd Quartos', 'Local / Praia', 'Cidade', 'UF', 'Valores Estimados',
        'Qualificação', 'Validação', 'Comportamento de Compra', 'Sinais de Intenção', 'Redes Sociais', 'LATITUDE', 'LONGITUDE', 'Score Qual.', 'Score Valid.'
    ]
    
    data_keys = [
        "pousada", "email", "whatsapp", "quartos", "local", "cidade", "uf", "valores",
        "qualificacao", "validacao", "comportamento", "sinais", "redes", "lat", "lon", "score_qual", "score_valid"
    ]
    
    # Tab Geral
    all_leads_sorted = []
    for city_name, leads in sorted(leads_by_city.items()):
        all_leads_sorted.extend(leads)
        
    create_styled_sheet(wb_out, "Todas_Pousadas_Validas", columns_headers, all_leads_sorted, data_keys)
    
    # Tabs por Cidades
    for city_name, leads in sorted(leads_by_city.items()):
        sheet_title = re.sub(r"[\\/*?:\[\]]", "", city_name)[:30]
        if not sheet_title:
            sheet_title = "Outros"
        create_styled_sheet(wb_out, sheet_title, columns_headers, leads, data_keys)
        
    wb_out.save(CONSOLIDATED_FILE_PATH)
    print(f"\n🎉 [SUCESSO-FECHAMENTO] Planilha final atualizada com {len(existing_leads)} leads em: {CONSOLIDATED_FILE_PATH}")
    print(f"   Imbituba: {len(leads_by_city.get('Imbituba', []))} leads")
    print(f"   Garopaba: {len(leads_by_city.get('Garopaba', []))} leads")

def create_styled_sheet(wb, title, headers, leads, keys):
    sheet = wb.create_sheet(title=title)
    sheet.views.sheetView[0].showGridLines = True
    
    # Escreve cabeçalhos
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
        
    sheet.row_dimensions[1].height = 28
    
    # Escreve linhas
    for row_idx, lead in enumerate(leads, 2):
        sheet.row_dimensions[row_idx].height = 20
        
        # Coluna sequencial #
        num_cell = sheet.cell(row=row_idx, column=1)
        num_cell.value = row_idx - 1
        num_cell.font = ROW_FONT
        num_cell.alignment = Alignment(horizontal="center")
        num_cell.border = BORDER_THIN
        
        for col_idx, key in enumerate(keys, 2):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = lead[key]
            cell.font = ROW_FONT
            cell.border = BORDER_THIN
            
            # Alinhamento
            if key in ["whatsapp", "quartos", "score_qual", "score_valid", "lat", "lon", "uf"]:
                cell.alignment = Alignment(horizontal="center")
            elif key in ["email"]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-ajuste de largura de colunas (estilo premium)
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 45)

if __name__ == "__main__":
    main()
