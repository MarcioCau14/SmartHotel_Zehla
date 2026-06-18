import openpyxl
import re

CONSOLIDATED_FILE_PATH = "/Users/marciocau/Downloads/PLANILHAS_MARKETING_BR_/CONTATOS_VALIDOS_CONSOLIDADO.xlsx"

def audit_duplicates():
    print("🕵️ [AUDIT] Carregando planilha consolidada...")
    wb = openpyxl.load_workbook(CONSOLIDATED_FILE_PATH, data_only=True)
    
    if "Todas_Pousadas_Validas" not in wb.sheetnames:
        print("❌ Aba 'Todas_Pousadas_Validas' não encontrada!")
        return
        
    sheet = wb["Todas_Pousadas_Validas"]
    print(f"📊 Total de linhas detectadas: {sheet.max_row}")
    
    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
    cols = {h: idx for idx, h in enumerate(headers, 1)}
    
    p_col = cols.get("pousada") or cols.get("nome")
    e_col = cols.get("e-mail") or cols.get("email")
    w_col = cols.get("whatsapp") or cols.get("whats")
    c_col = cols.get("cidade")
    
    if not p_col or not e_col or not w_col:
        print(f"❌ Colunas essenciais não identificadas! Headers: {headers}")
        return
        
    seen_names = {}
    seen_emails = {}
    seen_whatsapps = {}
    
    dup_names = []
    dup_emails = []
    dup_whatsapps = []
    
    for row in range(2, sheet.max_row + 1):
        name = str(sheet.cell(row=row, column=p_col).value or '').strip()
        email = str(sheet.cell(row=row, column=e_col).value or '').strip().lower()
        whats = str(sheet.cell(row=row, column=w_col).value or '').strip()
        cidade = str(sheet.cell(row=row, column=c_col).value or '').strip() if c_col else "-"
        
        if not name:
            continue
            
        name_norm = name.lower()
        whats_digits = re.sub(r"\D", "", whats)
        
        # Check Name
        if name_norm in seen_names:
            dup_names.append((row, name, cidade, seen_names[name_norm]))
        else:
            seen_names[name_norm] = (row, cidade)
            
        # Check Email
        if email and email != "-" and "@" in email:
            if email in seen_emails:
                dup_emails.append((row, email, name, seen_emails[email]))
            else:
                seen_emails[email] = (row, name)
                
        # Check WhatsApp
        if whats_digits:
            if whats_digits in seen_whatsapps:
                dup_whatsapps.append((row, whats, name, seen_whatsapps[whats_digits]))
            else:
                seen_whatsapps[whats_digits] = (row, name)
                
    print("\n================ AUDITORIA DE DUPLICATAS ================")
    print(f"Total de Pousadas Analisadas: {len(seen_names)}")
    print(f"Duplicatas de Nome: {len(dup_names)}")
    for dup in dup_names:
        print(f"  - Linha {dup[0]}: '{dup[1]}' em '{dup[2]}' duplicou com Linha {dup[3][0]} em '{dup[3][1]}'")
        
    print(f"Duplicatas de E-mail: {len(dup_emails)}")
    for dup in dup_emails:
        print(f"  - Linha {dup[0]}: '{dup[1]}' de '{dup[2]}' duplicou com Linha {dup[3][0]} de '{dup[3][1]}'")
        
    print(f"Duplicatas de WhatsApp: {len(dup_whatsapps)}")
    for dup in dup_whatsapps:
        print(f"  - Linha {dup[0]}: '{dup[1]}' de '{dup[2]}' duplicou com Linha {dup[3][0]} de '{dup[3][1]}'")
        
    if len(dup_names) == 0 and len(dup_emails) == 0 and len(dup_whatsapps) == 0:
        print("\n✅ INTEGRIDADE CONFIRMADA: 0 duplicatas globais encontradas!")
    else:
        print("\n⚠️ AVISO: Foram encontradas duplicatas! Verifique os logs acima.")

if __name__ == "__main__":
    audit_duplicates()
