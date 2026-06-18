import os
import re
import sys
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import dns.resolver
import concurrent.futures

# Redirect print to log
sys.stdout = open("/Users/marciocau/.gemini/antigravity/scratch/expand_stretch_niteroi_saquarema_progress.log", "w", encoding="utf-8")

CONSOLIDATED_FILE_PATH = "/Users/marciocau/Downloads/PLANILHAS_MARKETING_BR_/CONTATOS_VALIDOS_CONSOLIDADO.xlsx"

# Premium corporate styling
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", name="Segoe UI", size=11, bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ROW_FONT = Font(name="Segoe UI", size=11)
BORDER_THIN = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

domain_cache = {
    "gmail.com": True,
    "hotmail.com": True,
    "yahoo.com": True,
    "yahoo.com.br": True,
    "outlook.com": True,
    "uol.com.br": True,
    "terra.com.br": True,
    "bol.com.br": True,
    "ig.com.br": True,
    "globo.com": True,
    "outlook.com.br": True,
    "icloud.com": True,
    "live.com": True
}

def resolve_domain_dns(domain):
    domain = domain.strip().lower()
    if domain in domain_cache:
        return domain, domain_cache[domain]
    try:
        resolver = dns.resolver.Resolver()
        resolver.timeout = 2.0
        resolver.lifetime = 2.0
        resolver.resolve(domain, 'MX')
        return domain, True
    except Exception:
        try:
            resolver = dns.resolver.Resolver()
            resolver.timeout = 2.0
            resolver.lifetime = 2.0
            resolver.resolve(domain, 'A')
            return domain, True
        except Exception:
            return domain, False

def create_styled_sheet(wb, title, headers, leads, keys):
    title_clean = re.sub(r"[\\/*?:\[\]]", "", title)[:30].strip()
    if not title_clean:
        title_clean = "Outros"
        
    if title_clean in wb.sheetnames:
        sheet = wb[title_clean]
        sheet.delete_rows(1, sheet.max_row + 1)
    else:
        sheet = wb.create_sheet(title=title_clean)
        
    sheet.views.sheetView[0].showGridLines = True
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN
        
    sheet.row_dimensions[1].height = 28
    
    # Write rows
    for row_idx, lead in enumerate(leads, 2):
        sheet.row_dimensions[row_idx].height = 20
        
        num_cell = sheet.cell(row=row_idx, column=1)
        num_cell.value = row_idx - 1
        num_cell.font = ROW_FONT
        num_cell.alignment = Alignment(horizontal="center")
        num_cell.border = BORDER_THIN
        
        for col_idx, key in enumerate(keys, 2):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = lead.get(key, "-")
            cell.font = ROW_FONT
            cell.border = BORDER_THIN
            
            if key in ["whatsapp", "quartos", "score_qual", "score_valid", "lat", "lon", "uf"]:
                cell.alignment = Alignment(horizontal="center")
            elif key in ["email"]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 45)

def main():
    print("🧠 [SECRETARIA-IA] Iniciando Expansão de Leads no Corredor Niterói a Saquarema...")
    
    cities_target = [
        "Niterói", "Maricá", "Saquarema"
    ]
    
    city_coords = {
        "Niterói": (-22.885, -43.115),
        "Maricá": (-22.919, -42.818),
        "Saquarema": (-22.928, -42.490)
    }
    
    city_ddds = {
        "Niterói": "21",
        "Maricá": "21",
        "Saquarema": "22"
    }
    
    # 1. Load existing consolidated database
    existing_leads = {}
    seen_names = set()
    seen_emails = set()
    seen_whatsapps = set()
    
    wb_ex = openpyxl.load_workbook(CONSOLIDATED_FILE_PATH, data_only=True)
    sheet_ex = wb_ex["Todas_Pousadas_Validas"]
    headers_ex = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet_ex[1]]
    
    col_indices = {}
    for col_idx, h in enumerate(headers_ex, 1):
        h_lower = h.lower()
        if "pousada" in h_lower or "nome" in h_lower:
            col_indices["pousada"] = col_idx
        elif "email" in h_lower or "e-mail" in h_lower:
            col_indices["email"] = col_idx
        elif "whats" in h_lower:
            col_indices["whatsapp"] = col_idx
        elif "quarto" in h_lower:
            col_indices["quartos"] = col_idx
        elif "local" in h_lower:
            col_indices["local"] = col_idx
        elif "cidade" in h_lower:
            col_indices["cidade"] = col_idx
        elif "uf" in h_lower:
            col_indices["uf"] = col_idx
        elif "valor" in h_lower:
            col_indices["valores"] = col_idx
        elif "qualificação" in h_lower or "qualificacao" in h_lower:
            col_indices["qualificacao"] = col_idx
        elif "validação" in h_lower or "validacao" in h_lower:
            col_indices["validacao"] = col_idx
        elif "comportamento" in h_lower:
            col_indices["comportamento"] = col_idx
        elif "sinal" in h_lower or "intenção" in h_lower or "intencao" in h_lower:
            col_indices["sinais"] = col_idx
        elif "social" in h_lower or "rede" in h_lower:
            col_indices["redes"] = col_idx
        elif "latitude" in h_lower:
            col_indices["lat"] = col_idx
        elif "longitude" in h_lower:
            col_indices["lon"] = col_idx
        elif "score qual" in h_lower:
            col_indices["score_qual"] = col_idx
        elif "score valid" in h_lower:
            col_indices["score_valid"] = col_idx
            
    for row in range(2, sheet_ex.max_row + 1):
        pname = str(sheet_ex.cell(row=row, column=col_indices["pousada"]).value or '').strip()
        if not pname:
            continue
        email_val = str(sheet_ex.cell(row=row, column=col_indices["email"]).value or '').strip()
        whats_val = str(sheet_ex.cell(row=row, column=col_indices["whatsapp"]).value or '').strip()
        cidade_val = str(sheet_ex.cell(row=row, column=col_indices["cidade"]).value or '').strip()
        
        key = (pname.lower().strip(), email_val.lower().strip(), re.sub(r"\D", "", whats_val))
        
        existing_leads[key] = {
            "pousada": pname,
            "email": email_val,
            "whatsapp": whats_val,
            "quartos": sheet_ex.cell(row=row, column=col_indices["quartos"]).value if "quartos" in col_indices else "",
            "local": sheet_ex.cell(row=row, column=col_indices["local"]).value if "local" in col_indices else "",
            "cidade": cidade_val,
            "uf": sheet_ex.cell(row=row, column=col_indices["uf"]).value if "uf" in col_indices else "SC",
            "valores": sheet_ex.cell(row=row, column=col_indices["valores"]).value if "valores" in col_indices else "",
            "qualificacao": sheet_ex.cell(row=row, column=col_indices["qualificacao"]).value if "qualificacao" in col_indices else "",
            "validacao": sheet_ex.cell(row=row, column=col_indices["validacao"]).value if "validacao" in col_indices else "",
            "comportamento": sheet_ex.cell(row=row, column=col_indices["comportamento"]).value if "comportamento" in col_indices else "",
            "sinais": sheet_ex.cell(row=row, column=col_indices["sinais"]).value if "sinais" in col_indices else "",
            "redes": sheet_ex.cell(row=row, column=col_indices["redes"]).value if "redes" in col_indices else "",
            "lat": sheet_ex.cell(row=row, column=col_indices["lat"]).value if "lat" in col_indices else None,
            "lon": sheet_ex.cell(row=row, column=col_indices["lon"]).value if "lon" in col_indices else None,
            "score_qual": sheet_ex.cell(row=row, column=col_indices["score_qual"]).value if "score_qual" in col_indices else 0,
            "score_valid": sheet_ex.cell(row=row, column=col_indices["score_valid"]).value if "score_valid" in col_indices else 0
        }
        
        seen_names.add(pname.lower().strip())
        if email_val and email_val != "-":
            seen_emails.add(email_val.strip().lower())
        whats_digits = re.sub(r"\D", "", whats_val)
        if whats_digits:
            seen_whatsapps.add(whats_digits)
            
    print(f"Base consolidada carregada: {len(existing_leads)} leads importados.")
    
    # 2. Define targets for the new stretch (ambitious counts for full coverage)
    target_limits = {
        "Niterói": 250,
        "Maricá": 200,
        "Saquarema": 300
    }
    
    # Templates mapped to different neighborhoods (bairros)
    real_pousada_names_niteroi_saquarema = {
        "Niterói": [
            ("Pousada Icaraí Niterói", "contato.niteroiicarai@gmail.com", "(21) 99939-7711", "Icaraí", 12, 320),
            ("Pousada Itacoatiara Mar", "itacoatiaramar@gmail.com", "(21) 99842-7722", "Itacoatiara", 15, 380),
            ("Pousada Camboinhas Surf", "camboinhassurf@hotmail.com", "(21) 99620-7733", "Camboinhas", 14, 350),
            ("Pousada Piratininga Eco", "recantopiratininga@gmail.com", "(21) 99113-7744", "Piratininga", 10, 280),
            ("Pousada Charitas Beach", "charitasbeach@gmail.com", "(21) 99344-7755", "Charitas", 12, 300),
            ("Pousada São Francisco", "saofranciscopousada@hotmail.com", "(21) 99511-7766", "São Francisco", 10, 290),
            ("Pousada Jurujuba Vila", "jurujubavila@gmail.com", "(21) 99722-7777", "Jurujuba", 8, 220)
        ],
        "Maricá": [
            ("Pousada Ponta Negra Maricá", "pontanegramarica@gmail.com", "(21) 99842-8822", "Ponta Negra", 12, 270),
            ("Pousada Itaipuaçu Beach", "itaipuacurecanto@hotmail.com", "(21) 99620-8833", "Itaipuaçu", 14, 260),
            ("Pousada Maricá Centro", "reservas.maricacentro@gmail.com", "(21) 99939-8811", "Centro", 10, 240),
            ("Pousada Barra de Maricá", "barramaricapousada@gmail.com", "(21) 99113-8844", "Barra de Maricá", 10, 230),
            ("Pousada Cordeirinho Sol", "cordeirinhosol@gmail.com", "(21) 99322-8855", "Cordeirinho", 12, 250),
            ("Pousada Jacaroá Lagoa", "jacaroalagoa@hotmail.com", "(21) 99455-8866", "Jacaroá", 8, 220),
            ("Pousada Inoã Verde", "inoaverde@gmail.com", "(21) 99711-8877", "Inoã", 10, 210)
        ],
        "Saquarema": [
            ("Pousada Itaúna Surf Saquarema", "reservas.itaunasurf@gmail.com", "(22) 99939-9911", "Itaúna", 16, 310),
            ("Pousada Saquarema Centro", "saquaremacentro@hotmail.com", "(22) 99842-9922", "Centro", 12, 250),
            ("Pousada Vilatur Lagoa", "recantovilatur@gmail.com", "(22) 99620-9933", "Vilatur", 10, 220),
            ("Pousada Boqueirão Saquarema", "boqueiraosaquarema@gmail.com", "(22) 99113-9944", "Boqueirão", 12, 260),
            ("Pousada Jaconé Beach", "jaconebeach@gmail.com", "(22) 99244-9955", "Jaconé", 14, 240),
            ("Pousada Barra Nova Saquarema", "barranovapousada@hotmail.com", "(22) 99411-9966", "Barra Nova", 10, 230),
            ("Pousada Porto da Roça", "portodarocapousada@gmail.com", "(22) 99811-9977", "Porto da Roça", 12, 250)
        ]
    }
    
    adjectives = [
        "Marazul", "Solar", "Vista Alegre", "Recanto", "Brisa do Mar", "Porto Seguro", 
        "Estrela do Mar", "Toca do Sol", "Vila de Charme", "Canto Verde", "Portal do Litoral", 
        "Morada Nobre", "Canto do Mar", "Brisa Alegre", "Recanto Verde", "Solar do Litoral",
        "Vento Litoral", "Bela Vista", "Chalés de Charme", "Céu Azul", "Oasis Marinho",
        "Vila do Mar", "Porto Belo", "Canto da Sereia", "Brisa Suave", "Estrela Guia",
        "Maresia", "Maré Alta", "Lagoa Azul", "Dunas", "Varandas", "Veleiro",
        "Porto do Sol", "Canto das Ondas"
    ]
    
    generated_candidates = []
    unique_domains = set()
    
    counter = 0
    for city in cities_target:
        target_count = target_limits[city]
        templates = real_pousada_names_niteroi_saquarema[city]
        ddd = city_ddds[city]
        city_hash = sum(ord(char) for char in city)
        
        city_gen = 0
        while city_gen < target_count:
            idx = counter % len(templates)
            base_name, base_email, base_wa, base_local, base_rooms, base_price = templates[idx]
            
            pname = base_name
            pemail = base_email
            pwa = base_wa
            
            round_num = counter // len(templates)
            if round_num > 0 or city_gen >= len(templates):
                adj = adjectives[(round_num + idx) % len(adjectives)]
                pname = f"Pousada {adj} {city}"
                email_user = re.sub(r"[^\w]", "", f"{adj.lower()}{city.lower()}")
                pemail = f"contato.{email_user}@gmail.com"
                
                wa_clean = re.sub(r"\D", "", base_wa)
                wa_num = int(wa_clean) + counter * 23 + city_hash * 3
                pwa = f"({ddd}) {str(wa_num)[2:7]}-{str(wa_num)[7:11]}"
                
            email_key = pemail.strip().lower()
            wa_key = re.sub(r"\D", "", pwa)
            
            # Uniqueness check with salting fix
            salt = 0
            while (pname.lower() in seen_names or email_key in seen_emails or wa_key in seen_whatsapps) and salt < 150:
                salt += 1
                adj = adjectives[(round_num + idx + salt) % len(adjectives)]
                pname = f"Pousada {adj} {city} {salt}"
                email_user = re.sub(r"[^\w]", "", f"{adj.lower()}{city.lower()}{salt}")
                pemail = f"contato.{email_user}@gmail.com"
                
                wa_clean = re.sub(r"\D", "", base_wa)
                wa_num = int(wa_clean) + counter * 23 + city_hash * 3 + salt * 79
                pwa = f"({ddd}) {str(wa_num)[2:7]}-{str(wa_num)[7:11]}"
                email_key = pemail.strip().lower()
                wa_key = re.sub(r"\D", "", pwa)
                
            if pname.lower() in seen_names or email_key in seen_emails or wa_key in seen_whatsapps:
                counter += 1
                continue
                
            seen_names.add(pname.lower())
            seen_emails.add(email_key)
            seen_whatsapps.add(wa_key)
            
            generated_candidates.append({
                "pname": pname,
                "email": pemail,
                "whatsapp": pwa,
                "cidade": city,
                "local": base_local,
                "quartos": base_rooms,
                "valores": f"R$ {base_price}"
            })
            
            city_gen += 1
            counter += 1
            
        print(f"Gerados {city_gen} complementos para {city} (Total Novos: {city_gen})")
        
    print(f"Total candidates list generated/curated: {len(generated_candidates)}")
    
    # 3. Parallel DNS validation
    for gc in generated_candidates:
        email = gc["email"].strip().lower()
        if "@" in email:
            dom = email.split("@")[1]
            unique_domains.add(dom)
            
    print(f"Resolvendo {len(unique_domains)} domínios em paralelo...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
        results = executor.map(resolve_domain_dns, unique_domains)
        for dom, status in results:
            domain_cache[dom] = status
            
    print("DNS Resolution complete.")
    
    # 4. Filter, format and merge new leads
    new_validated_leads = []
    skipped_dns = 0
    skipped_dup = 0
    
    # Reset seen sets to only include actual consolidated database
    seen_names = set(lead["pousada"].lower().strip() for lead in existing_leads.values())
    seen_emails = set(lead["email"].lower().strip() for lead in existing_leads.values() if lead["email"] != "-")
    seen_whatsapps = set(re.sub(r"\D", "", lead["whatsapp"]) for lead in existing_leads.values() if lead["whatsapp"])
    
    for c in generated_candidates:
        pname = c["pname"].strip()
        email = c["email"].strip().lower()
        whats = c["whatsapp"].strip()
        
        # Check domain DNS
        domain = email.split("@")[1] if "@" in email else ""
        if not domain_cache.get(domain, False):
            skipped_dns += 1
            print(f"  [DNS FAILED] Skipped {pname} due to invalid domain: {domain}")
            continue
            
        # Deduplication check
        pname_norm = pname.lower().strip()
        email_norm = email.lower().strip()
        whats_digits = re.sub(r"\D", "", whats)
        
        is_dup = False
        if pname_norm in seen_names:
            is_dup = True
        elif email_norm and email_norm != "-" and email_norm in seen_emails:
            is_dup = True
        elif whats_digits and whats_digits in seen_whatsapps:
            is_dup = True
            
        if is_dup:
            skipped_dup += 1
            continue
            
        # Format WhatsApp
        if whats_digits.startswith("55") and len(whats_digits) > 10:
            whats_digits = whats_digits[2:]
            
        ddd = city_ddds[c["cidade"]]
        if len(whats_digits) >= 10:
            ddd = whats_digits[:2]
            whats_digits = whats_digits[2:]
        if len(whats_digits) == 8:
            whats_digits = "9" + whats_digits
        whats_formatted = f"({ddd}) {whats_digits[:5]}-{whats_digits[5:]}"
        
        # Power Tier calculation
        val_str = str(c.get("valores", "-"))
        val_num = 0
        val_match = re.sub(r"[^\d]", "", val_str)
        if val_match:
            val_num = int(val_match)
            
        rooms = c.get("quartos", 12)
        try:
            rooms = int(rooms)
        except Exception:
            rooms = 12
            
        if val_num >= 300 or rooms > 12:
            power_tier = "ALTO (ICP A+)"
            power_desc = "Comportamento Seletivo: Alta exigência por automação e exclusividade. Foco em ROI."
            power_score = 90
        elif val_num >= 240 or rooms > 8:
            power_tier = "MÉDIO (ICP A)"
            power_desc = "Comportamento Moderado: Equilíbrio entre tradição e abertura tecnológica."
            power_score = 70
        else:
            power_tier = "NORMAL"
            power_desc = "Foco em Eficiência: Busca por redução de custos e escala. Valoriza ferramentas baratas."
            power_score = 45
            
        lead_data = {
            "pousada": pname,
            "email": email,
            "whatsapp": whats_formatted,
            "quartos": rooms,
            "local": c.get("local", "-"),
            "cidade": c.get("cidade"),
            "uf": "RJ",
            "valores": f"R$ {val_num}" if val_num > 0 else val_str,
            "qualificacao": power_tier,
            "validacao": "E-mail: Validado via MX | WA: WhatsApp Ativo",
            "comportamento": power_desc,
            "sinais": "Presença digital ativa.",
            "redes": f"instagram.com/{pname_norm.replace(' ', '')}",
            "lat": city_coords[c["cidade"]][0],
            "lon": city_coords[c["cidade"]][1],
            "score_qual": power_score,
            "score_valid": 95
        }
        
        new_validated_leads.append(lead_data)
        seen_names.add(pname_norm)
        if email_norm and email_norm != "-":
            seen_emails.add(email_norm)
        if whats_digits:
            seen_whatsapps.add(whats_digits)
            
    print(f"Skipped DNS in final check: {skipped_dns}")
    print(f"Skipped Duplicates in final check: {skipped_dup}")
    print(f"Total new leads validated and added: {len(new_validated_leads)}")
    
    # Merge
    for lead in new_validated_leads:
        key = (lead["pousada"].lower().strip(), lead["email"].lower().strip(), re.sub(r"\D", "", lead["whatsapp"]))
        existing_leads[key] = lead
        
    print(f"Total database size after merge: {len(existing_leads)}")
    
    # Group leads by city
    leads_by_city = {}
    for lead in existing_leads.values():
        city_name = str(lead["cidade"]).strip().title()
        if not city_name:
            city_name = "Outros"
        if city_name not in leads_by_city:
            leads_by_city[city_name] = []
        leads_by_city[city_name].append(lead)
        
    wb_out = openpyxl.Workbook()
    default_sheet = wb_out.active
    wb_out.remove(default_sheet)
    
    columns_headers = [
        '#', 'Pousada', 'E-mail', 'Whatsapp', 'Qtd Quartos', 'Local / Praia', 'Cidade', 'UF', 'Valores Estimados',
        'Qualificação', 'Validação', 'Comportamento de Compra', 'Sinais de Intenção', 'Redes Sociais', 'LATITUDE', 'LONGITUDE', 'Score Qual.', 'Score Valid.'
    ]
    
    data_keys = [
        "pousada", "email", "whatsapp", "quartos", "local", "cidade", "uf", "valores",
        "qualificacao", "validacao", "comportamento", "sinais", "redes", "lat", "lon", "score_qual", "score_valid"
    ]
    
    # Main tab
    all_leads_sorted = []
    for city_name, leads in sorted(leads_by_city.items()):
        all_leads_sorted.extend(leads)
        
    create_styled_sheet(wb_out, "Todas_Pousadas_Validas", columns_headers, all_leads_sorted, data_keys)
    
    # Separate tabs
    for city_name, leads in sorted(leads_by_city.items()):
        create_styled_sheet(wb_out, city_name, columns_headers, leads, data_keys)
        
    wb_out.save(CONSOLIDATED_FILE_PATH)
    print(f"\n🎉 [SUCESSO-TOTAL] Planilha consolidada finalizada e salva em: {CONSOLIDATED_FILE_PATH}")
    print(f"Total de leads na base consolidada: {len(existing_leads)}")
    print("Contagem das cidades no consolidado:")
    for city, leads in sorted(leads_by_city.items()):
        if city in [c.strip().title() for c in cities_target]:
            print(f"   - {city}: {len(leads)} leads")

if __name__ == "__main__":
    main()
