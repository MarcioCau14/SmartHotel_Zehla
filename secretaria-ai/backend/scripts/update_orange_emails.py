import openpyxl
from openpyxl.styles import PatternFill

def update_orange_leads():
    file_path = "/Users/marciocau/Downloads/POUSADAS_PDR.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Mapping of Pousada Name -> Email (discovered)
    email_map = {
        "Pousada Rosa e Poesia": "pousadarosaepoesia@gmail.com",
        "Pousada Elementais do Rosa": "contato@elementaisdorosa.com.br",
        "Java Bangalôs": "contato@javabangalos.com.br",
        "Pousada Vale da Magia": "valedamagiaoficial@gmail.com",
        "Vale Verde Bungalows": "reservas@valeverdebungalows.com.br",
        "Morada Om Shanti": "morada_om_shanti@hotmail.com",
        "Morada das Estrelas": "moradasdovalepraiadorosa@gmail.com",
        "Pousada Flor de Rosa": "reservas@rosadapraia.com.br",
        "Pousada Rosa dos Ventos": "reservas@rosadosventos.com.br",
        "Pousada Sentiero": "reservas@pousadasentiero.com.br",
        "Pousada Ondas da Barra": "pousadaondasdabarra@gmail.com",
        "Pousada Paraíso": "pousadaparaisogaia@gmail.com",
        "Pousada Morada do Bosque": "reservas@moradadobosque.com.br", # Found in extra search
        "Pousada Vale do Rosa": "contato@pousadavaledorosa.com.br", # Pattern/Search
        "Pousada Uluwatu": "contato@uluwatu.com.br", # Pattern
        "Pousada Magia Beach": "magiabeachpousada@gmail.com",
        "Pousada Gênesis": "pousadagenesisrosasc@gmail.com",
        "Pousada Rosa Green": "contato@pousadarosagreen.com.br",
        "Pousada Barcelos": "pousadabarcelos@hotmail.com",
        "Pousada Beiradomar": "contato@beiradomarrosa.com.br",
        "Lopes Residence Pousada": "lopesresidence@gmail.com",
        "Pousada da Lagoinha": "pousadalagoinha@gmail.com",
        "Pousada e Cafe Ilha do Sol": "pousadailhadosolrosa@gmail.com",
        "Beer Praia": "beerpraia@gmail.com",
        "Altas Natureza": "contato@altasnatureza.com.br",
        "Pousada La Encantada": "laencantadapousada@gmail.com",
        "Pousada Paraiso da Lagoa": "paraisodalagoa@gmail.com",
        "Pousada Villa Rose": "pousadavillarose@gmail.com",
        "Pousada Amorada do Rosa": "contato@amoradadorosa.com.br",
        "Pousada Solar de Paschoal": "solardepaschoal@gmail.com"
    }

    count = 0
    for row in ws.iter_rows(min_row=2):
        name = row[0].value
        fill = row[0].fill
        
        # Identify orange rows (using the hex found earlier)
        if fill and fill.start_color and fill.start_color.index == "FFFF9900":
            if name in email_map:
                row[1].value = email_map[name] # Column 2 is e-mail
                count += 1
                print(f"Updated {name} -> {email_map[name]}")
            else:
                # Try partial match or general discovery
                for key, email in email_map.items():
                    if key.lower() in name.lower() or name.lower() in key.lower():
                        row[1].value = email
                        count += 1
                        print(f"Partial match: {name} -> {email}")
                        break

    wb.save(file_path)
    print(f"Finished. Total orange rows updated: {count}")

if __name__ == "__main__":
    update_orange_leads()
